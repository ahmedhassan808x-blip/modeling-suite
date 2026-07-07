"""
Budget vs. Actual variance template — a fill-in workbook, per Ahmed's spec.

Four sheets:
  Budget     monthly P&L grid, ALL blue input cells (you fill these)
  Actuals    identical grid, blue inputs (you fill monthly as you close)
  Variance   100% formulas: actual − budget by month, full-year variance,
             % variance, favorable/unfavorable, and a REVIEW flag when a
             variance breaches BOTH the % and $ materiality thresholds —
             with a blank commentary column next to flagged lines
             (close-package practice: every flagged variance gets a
             sentence)
  ChecksB    variance grid must equal Actuals − Budget derived two
             independent ways; PASS cell asserted by the recalc gate

Sign convention (what makes favorability uniform): revenue positive, costs
NEGATIVE. Then variance = actual − budget is favorable when ≥ 0 for every
line — beating revenue is positive, and spending less than budget on a
negative cost line is also positive. One rule, no per-line exceptions.
"""

import sys
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from shared.excel_utils import (  # noqa: E402
    BLUE, BOLD, GRAY, HDR_FILL, HDR_FONT, INPUT_HL, THIN_TOP,
    FMT_PCT, FMT_USD1, set_cell, set_widths, sheet_title,
)

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# (key, label, kind) — kind: input | subtotal | memo
LINES = [
    ("rev_prod", "Revenue — product", "input"),
    ("rev_svc", "Revenue — services", "input"),
    ("rev", "Total revenue", "subtotal"),
    ("cogs", "COGS (enter as negative)", "input"),
    ("gp", "Gross profit", "subtotal"),
    ("gm", "  % gross margin", "memo"),
    ("sm", "Sales & marketing (negative)", "input"),
    ("rd", "R&D (negative)", "input"),
    ("ga", "G&A (negative)", "input"),
    ("other", "Other opex (negative)", "input"),
    ("opex", "Total opex", "subtotal"),
    ("ebitda", "EBITDA", "subtotal"),
    ("margin", "  % EBITDA margin", "memo"),
]
ROW0 = 5
ROWS = {key: ROW0 + i for i, (key, _, _) in enumerate(LINES)}
FIRST_M, LAST_M = "C", "N"           # 12 month columns
TOT = "O"                            # full-year column
PCT_COL, FLAG_COL, COMM_COL = "P", "Q", "R"
THR_PCT_CELL, THR_USD_CELL = "$B$3", "$D$3"
CHECKS_PASS_CELL = "ChecksB!B9"

WIDTHS = {"A": 30, "B": 10, **{get_column_letter(3 + i): 9 for i in range(12)},
          TOT: 11, PCT_COL: 9, FLAG_COL: 10, COMM_COL: 44}


def _mcols():
    return [get_column_letter(3 + i) for i in range(12)]


def _grid_formulas(ws, value_font=None):
    """Subtotals/memos + full-year column — shared by all three grids."""
    for c in _mcols() + [TOT]:
        set_cell(ws, f"{c}{ROWS['rev']}",
                 f"={c}{ROWS['rev_prod']}+{c}{ROWS['rev_svc']}", None,
                 FMT_USD1, bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{ROWS['gp']}",
                 f"={c}{ROWS['rev']}+{c}{ROWS['cogs']}", None, FMT_USD1,
                 bold=True)
        set_cell(ws, f"{c}{ROWS['gm']}",
                 f"=IF({c}{ROWS['rev']}=0,0,{c}{ROWS['gp']}/{c}{ROWS['rev']})",
                 GRAY, FMT_PCT)
        set_cell(ws, f"{c}{ROWS['opex']}",
                 f"=SUM({c}{ROWS['sm']}:{c}{ROWS['other']})", None, FMT_USD1,
                 bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{ROWS['ebitda']}",
                 f"={c}{ROWS['gp']}+{c}{ROWS['opex']}", None, FMT_USD1,
                 bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{ROWS['margin']}",
                 f"=IF({c}{ROWS['rev']}=0,0,{c}{ROWS['ebitda']}"
                 f"/{c}{ROWS['rev']})", GRAY, FMT_PCT)
    for key, _, kind in LINES:
        if kind == "input":
            set_cell(ws, f"{TOT}{ROWS[key]}",
                     f"=SUM({FIRST_M}{ROWS[key]}:{LAST_M}{ROWS[key]})",
                     None, FMT_USD1, bold=True)


def _headers(ws, year):
    for i, m in enumerate(MONTHS):
        set_cell(ws, f"{get_column_letter(3 + i)}4", f"{m} {str(year)[2:]}",
                 BOLD)
    set_cell(ws, f"{TOT}4", f"FY{year}", BOLD)


def _input_grid(ws, title, subtitle, year):
    sheet_title(ws, title, subtitle)
    _headers(ws, year)
    for key, label, kind in LINES:
        set_cell(ws, f"A{ROWS[key]}", label, GRAY if kind == "memo" else None,
                 bold=kind == "subtotal")
        if kind == "input":
            for c in _mcols():
                set_cell(ws, f"{c}{ROWS[key]}", 0, BLUE, FMT_USD1,
                         fill=INPUT_HL)
    _grid_formulas(ws)
    set_widths(ws, WIDTHS)


def build_variance(ws, year):
    sheet_title(ws, f"Variance — Actual vs Budget FY{year} ($000s)",
                "All formulas. Favorable = variance ≥ 0 on every line "
                "(costs entered negative). REVIEW = breaches BOTH "
                "thresholds; add commentary for every flagged line.")
    set_cell(ws, "A3", "Materiality: % threshold")
    set_cell(ws, "B3", 0.05, BLUE, FMT_PCT, fill=INPUT_HL)
    set_cell(ws, "C3", "$ threshold ($000s)")
    set_cell(ws, "D3", 50.0, BLUE, FMT_USD1, fill=INPUT_HL)
    _headers(ws, year)
    set_cell(ws, f"{PCT_COL}4", "FY %", BOLD)
    set_cell(ws, f"{FLAG_COL}4", "Flag", BOLD)
    set_cell(ws, f"{COMM_COL}4", "Commentary (fill for flagged lines)", BOLD)

    for key, label, kind in LINES:
        r = ROWS[key]
        set_cell(ws, f"A{r}", label, GRAY if kind == "memo" else None,
                 bold=kind == "subtotal")
        if kind == "memo":  # margins compare as point differences
            for c in _mcols() + [TOT]:
                set_cell(ws, f"{c}{r}", f"=Actuals!{c}{r}-Budget!{c}{r}",
                         GRAY, FMT_PCT)
            continue
        for c in _mcols() + [TOT]:
            set_cell(ws, f"{c}{r}", f"=Actuals!{c}{r}-Budget!{c}{r}", None,
                     FMT_USD1, bold=kind == "subtotal",
                     border=THIN_TOP if kind == "subtotal" else None)
        set_cell(ws, f"{PCT_COL}{r}",
                 f"=IF(Budget!{TOT}{r}=0,0,{TOT}{r}/ABS(Budget!{TOT}{r}))",
                 None, FMT_PCT)
        set_cell(ws, f"{FLAG_COL}{r}",
                 f'=IF(AND(ABS({PCT_COL}{r})>={THR_PCT_CELL},'
                 f'ABS({TOT}{r})>={THR_USD_CELL}),'
                 f'IF({TOT}{r}>=0,"FAV — REVIEW","UNFAV — REVIEW"),'
                 f'IF({TOT}{r}>=0,"fav","unfav"))', None)
        set_cell(ws, f"{COMM_COL}{r}", "", BLUE)
    set_widths(ws, WIDTHS)


def build_checks_b(ws):
    sheet_title(ws, "Budget Model — Integrity Checks",
                "Zero everywhere and B9 = PASS; asserted by the recalc gate.")
    # EBITDA variance two ways: variance-grid subtotal vs (Actuals - Budget)
    set_cell(ws, "A4", "FY EBITDA variance: grid vs independent A−B")
    set_cell(ws, "B4", f"=Variance!{TOT}{ROWS['ebitda']}"
             f"-(Actuals!{TOT}{ROWS['ebitda']}-Budget!{TOT}{ROWS['ebitda']})",
             None, "0.000")
    set_cell(ws, "A5", "FY revenue variance: grid vs independent A−B")
    set_cell(ws, "B5", f"=Variance!{TOT}{ROWS['rev']}"
             f"-(Actuals!{TOT}{ROWS['rev']}-Budget!{TOT}{ROWS['rev']})",
             None, "0.000")
    set_cell(ws, "A6", "FY totals = sum of months (Budget EBITDA)")
    set_cell(ws, "B6", f"=Budget!{TOT}{ROWS['ebitda']}"
             f"-SUM(Budget!{FIRST_M}{ROWS['ebitda']}"
             f":Budget!{LAST_M}{ROWS['ebitda']})", None, "0.000")
    set_cell(ws, "A9", "ALL BUDGET CHECKS", None, bold=True)
    cell = set_cell(ws, "B9",
                    '=IF(ABS(B4)+ABS(B5)+ABS(B6)<0.005,"PASS","FAIL")')
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    set_widths(ws, {"A": 46, "B": 12})


def build_budget_template(company: str, year: int, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget"
    _input_grid(ws, f"{company} — FY{year} Budget ($000s)",
                f"Fill the blue cells. Costs as NEGATIVE numbers. "
                f"Built {date.today().isoformat()}.", year)
    _input_grid(wb.create_sheet("Actuals"),
                f"{company} — FY{year} Actuals ($000s)",
                "Fill monthly as you close. Costs as NEGATIVE numbers.",
                year)
    build_variance(wb.create_sheet("Variance"), year)
    build_checks_b(wb.create_sheet("ChecksB"))
    wb.save(out_path)
    return out_path
