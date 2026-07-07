"""
DCF workbook builder — extends the three-statement model.

Ported from analyst-toolkit/valuation-engine (WACC build-up, Gordon terminal
value, sensitivity grid, comps, football field) with the structural upgrade
that motivated this suite: unlevered FCF is no longer a simplified standalone
build — EBIT, D&A, capex and ΔNWC are green links into the fully linked
three-statement forecast. Flip the scenario toggle on Assumptions and the
DCF reprices live.

Extensions over the old engine (its own README's v2 list):
  - exit-multiple terminal value alongside Gordon growth (bankers show both)
  - actual net debt pulled from the modeled balance sheet, not a hardcode
"""

import sys
from pathlib import Path

from openpyxl.chart import BarChart, Reference

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from shared.excel_utils import (  # noqa: E402
    BLUE, BOLD, GRAY, HDR_FILL, HDR_FONT, INPUT_HL, NAVY, THIN_TOP,
    FMT_PCT, FMT_PS, FMT_USD, FMT_X, col_letter, set_cell, set_widths,
    sheet_title, year_headers,
)
from three_statement.model_builder import (  # noqa: E402
    ASM, BS, CF, IS, N_FC, WIDTHS, build_model as build_three_statement_wb,
)

D = dict(ebit=5, tax=6, nopat=7, da=8, capex=9, nwc=10, ufcf=11, df=13, pv=14,
         price=18, shares=19, beta=20, rf=21, erp=22, kd=23, dw=24, taxw=25,
         ke=26, wacc=27,
         sum_pv=30, tg=31, tv=32, pv_tv=33, ev=34, net_debt=35, eq=36, ps=37,
         cur=38, upside=39,
         exit_mult=42, tv_exit=43, pv_tv_exit=44, ev_exit=45, ps_exit=46,
         sens_hdr=49, sens_first=50)
SENS_COLS = list("CDEFG")


def derive_wacc_seeds(data, a, peers=None) -> dict:
    """Seed WACC/terminal inputs from data. Blue and challengeable in Excel."""
    bsl = data["bs"]
    net_debt = bsl["debt"][-1] - bsl["cash"][-1]
    mkt_cap = data["price"] * data["shares_mm"]
    dw = max(0.0, min(0.6, net_debt / (net_debt + mkt_cap))) \
        if mkt_cap and net_debt + mkt_cap > 0 else 0.10
    peer_ebitda = sorted(p["ev_ebitda"] for p in (peers or []) if p["ev_ebitda"])
    exit_mult = peer_ebitda[len(peer_ebitda) // 2] if peer_ebitda else 10.0
    return dict(rf=0.0425, erp=0.05, kd=a.rate_debt, dw=round(dw, 3),
                tg=0.025, exit_mult=round(exit_mult, 1))


def build_dcf_sheet(ws, data, w, n_hist):
    fc = [col_letter(n_hist + i) for i in range(N_FC)]
    last_hist = col_letter(n_hist - 1)
    sheet_title(ws, f"{data['ticker']} — DCF (FCFF, linked to 3-statement)",
                "UFCF pulls live from the statement forecast — the scenario "
                "toggle on Assumptions reprices this sheet. Interest is "
                "excluded (unlevered); taxes at the driver rate on EBIT.")
    year_headers(ws, 4, data["years"], N_FC)

    labels = [("ebit", "EBIT"), ("tax", "Unlevered taxes on EBIT"),
              ("nopat", "NOPAT"), ("da", "(+) D&A"), ("capex", "(–) Capex"),
              ("nwc", "(–) Increase in NWC"), ("ufcf", "Unlevered FCF"),
              ("df", "Discount factor"), ("pv", "PV of UFCF")]
    for key, label in labels:
        set_cell(ws, f"A{D[key]}", label, bold=key in ("nopat", "ufcf"))

    for t, c in enumerate(fc, start=1):
        set_cell(ws, f"{c}{D['ebit']}", f"=IS!{c}{IS['ebit']}", None, FMT_USD)
        set_cell(ws, f"{c}{D['tax']}",
                 f"=-{c}{D['ebit']}*Assumptions!{c}{ASM['tax']}", None, FMT_USD)
        set_cell(ws, f"{c}{D['nopat']}", f"={c}{D['ebit']}+{c}{D['tax']}",
                 None, FMT_USD, bold=True)
        set_cell(ws, f"{c}{D['da']}", f"=-IS!{c}{IS['da']}", None, FMT_USD)
        set_cell(ws, f"{c}{D['capex']}", f"=CF!{c}{CF['capex']}", None, FMT_USD)
        # NB: explicit addition, not SUM(CF!x:CF!y) — LibreOffice parses the
        # repeated-prefix range syntax differently and computes silently
        # wrong values (caught by the hand-checked UFCF test).
        nwc_refs = "+".join(f"CF!{c}{CF[k]}"
                            for k in ("d_ar", "d_inv", "d_oca", "d_ap", "d_ocl"))
        set_cell(ws, f"{c}{D['nwc']}", f"={nwc_refs}", None, FMT_USD)
        set_cell(ws, f"{c}{D['ufcf']}", f"=SUM({c}{D['nopat']}:{c}{D['nwc']})",
                 None, FMT_USD, bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{D['df']}", f"=1/(1+$B${D['wacc']})^{t}", None, "0.000")
        set_cell(ws, f"{c}{D['pv']}", f"={c}{D['ufcf']}*{c}{D['df']}",
                 None, FMT_USD)

    set_cell(ws, f"A{D['price'] - 1}", "WACC build-up", bold=True)
    wacc_rows = [
        ("price", "Share price ($)", data["price"], BLUE, FMT_PS),
        ("shares", "Shares outstanding (mm)", data["shares_mm"], BLUE, FMT_USD),
        ("beta", "Beta", data["beta"], BLUE, "0.00"),
        ("rf", "Risk-free rate (update to current 10Y)", w["rf"], BLUE, FMT_PCT),
        ("erp", "Equity risk premium", w["erp"], BLUE, FMT_PCT),
        ("kd", "Pre-tax cost of debt", w["kd"], BLUE, FMT_PCT),
        ("dw", "Debt / (Debt + Equity)", w["dw"], BLUE, FMT_PCT),
        ("taxw", "Tax rate (Y1 driver)",
         f"=Assumptions!{fc[0]}{ASM['tax']}", None, FMT_PCT),
        ("ke", "Cost of equity",
         f"=B{D['rf']}+B{D['beta']}*B{D['erp']}", None, FMT_PCT),
        ("wacc", "WACC",
         f"=(1-B{D['dw']})*B{D['ke']}+B{D['dw']}*B{D['kd']}*(1-B{D['taxw']})",
         None, FMT_PCT),
    ]
    for key, label, val, font, fmt in wacc_rows:
        set_cell(ws, f"A{D[key]}", label, bold=key == "wacc")
        cell = set_cell(ws, f"B{D[key]}", val, font, fmt, bold=key == "wacc")
        if key in ("rf", "erp", "kd", "dw"):
            cell.fill = INPUT_HL

    set_cell(ws, f"A{D['sum_pv'] - 1}", "Valuation — Gordon growth", bold=True)
    val_rows = [
        ("sum_pv", "Sum PV of UFCF ($mm)",
         f"=SUM({fc[0]}{D['pv']}:{fc[-1]}{D['pv']})", FMT_USD),
        ("tg", "Terminal growth (g)", w["tg"], FMT_PCT),
        ("tv", "Terminal value ($mm)",
         f"={fc[-1]}{D['ufcf']}*(1+B{D['tg']})/($B${D['wacc']}-B{D['tg']})",
         FMT_USD),
        ("pv_tv", "PV of terminal value",
         f"=B{D['tv']}/(1+$B${D['wacc']})^{N_FC}", FMT_USD),
        ("ev", "Enterprise value ($mm)", f"=B{D['sum_pv']}+B{D['pv_tv']}",
         FMT_USD),
        ("net_debt", "(–) Net debt (modeled BS, actual)",
         f"=BS!{last_hist}{BS['ltd']}+BS!{last_hist}{BS['rev']}"
         f"-BS!{last_hist}{BS['cash']}", FMT_USD),
        ("eq", "Equity value ($mm)", f"=B{D['ev']}-B{D['net_debt']}", FMT_USD),
        ("ps", "Implied value per share", f"=B{D['eq']}/B{D['shares']}", FMT_PS),
        ("cur", "Current price", f"=B{D['price']}", FMT_PS),
        ("upside", "Upside / (downside)", f"=B{D['ps']}/B{D['cur']}-1", FMT_PCT),
    ]
    for key, label, val, fmt in val_rows:
        bold = key in ("ev", "ps")
        set_cell(ws, f"A{D[key]}", label, bold=bold)
        cell = set_cell(ws, f"B{D[key]}", val, BLUE if key == "tg" else None,
                        fmt, bold=bold)
        if key == "tg":
            cell.fill = INPUT_HL

    set_cell(ws, f"A{D['exit_mult'] - 1}",
             "Terminal value cross-check — exit multiple", bold=True)
    exit_rows = [
        ("exit_mult", "Exit EV/EBITDA multiple"
         + (" (seed: peer median)" if w.get("exit_seeded") else ""),
         w["exit_mult"], BLUE, FMT_X),
        ("tv_exit", "Terminal value ($mm)",
         f"=B{D['exit_mult']}*IS!{fc[-1]}{IS['ebitda']}", None, FMT_USD),
        ("pv_tv_exit", "PV of terminal value",
         f"=B{D['tv_exit']}/(1+$B${D['wacc']})^{N_FC}", None, FMT_USD),
        ("ev_exit", "Enterprise value ($mm)",
         f"=B{D['sum_pv']}+B{D['pv_tv_exit']}", None, FMT_USD),
        ("ps_exit", "Implied value per share (exit method)",
         f"=(B{D['ev_exit']}-B{D['net_debt']})/B{D['shares']}", None, FMT_PS),
    ]
    for key, label, val, font, fmt in exit_rows:
        set_cell(ws, f"A{D[key]}", label, bold=key == "ps_exit")
        cell = set_cell(ws, f"B{D[key]}", val, font, fmt, bold=key == "ps_exit")
        if key == "exit_mult":
            cell.fill = INPUT_HL

    # Sensitivity: WACC (rows) x terminal growth (cols) -> implied $/share
    set_cell(ws, f"A{D['sens_hdr'] - 1}",
             "Sensitivity — implied value per share (Gordon)", bold=True)
    set_cell(ws, f"B{D['sens_hdr']}", "WACC \\ g", BOLD)
    for j, sc in enumerate(SENS_COLS):
        set_cell(ws, f"{sc}{D['sens_hdr']}",
                 f"=$B${D['tg']}+{j - 2}*0.005", None, FMT_PCT)
    for i in range(5):
        r = D["sens_first"] + i
        set_cell(ws, f"B{r}", f"=$B${D['wacc']}+{i - 2}*0.01", None, FMT_PCT)
        for sc in SENS_COLS:
            wref, gref = f"$B{r}", f"{sc}${D['sens_hdr']}"
            pv = "+".join(f"{c}${D['ufcf']}/(1+{wref})^{k + 1}"
                          for k, c in enumerate(fc))
            tv = (f"{fc[-1]}${D['ufcf']}*(1+{gref})/({wref}-{gref})"
                  f"/(1+{wref})^{N_FC}")
            set_cell(ws, f"{sc}{r}",
                     f"=(({pv}+{tv})-$B${D['net_debt']})/$B${D['shares']}",
                     None, FMT_PS)
    set_widths(ws, dict(WIDTHS, A=42))


def build_comps_sheet(ws, data, peers, n_hist):
    last_hist = col_letter(n_hist - 1)
    sheet_title(ws, f"{data['ticker']} — Trading Comparables",
                "Peer LTM multiples (blue, FMP snapshot) applied to the "
                "target's last reported metrics from the IS sheet.")
    headers = ["Ticker", "Company", "Mkt Cap ($mm)", "EV ($mm)",
               "EV/EBITDA", "EV/Revenue", "P/E"]
    for j, h in enumerate(headers):
        set_cell(ws, f"{col_letter(j, 1)}4", h, HDR_FONT, fill=HDR_FILL)

    first, last = 5, 4 + len(peers)
    for i, p in enumerate(peers):
        vals = [p["ticker"], p["name"], p["mkt_cap"], p["ev"],
                p["ev_ebitda"], p["ev_revenue"], p["pe"]]
        fmts = [None, None, FMT_USD, FMT_USD, FMT_X, FMT_X, FMT_X]
        for j, (v, f) in enumerate(zip(vals, fmts)):
            set_cell(ws, f"{col_letter(j, 1)}{first + i}", v, BLUE, f)

    stats = last + 2
    for k, label in enumerate(("Peer 25th pct", "Peer median", "Peer 75th pct")):
        set_cell(ws, f"B{stats + k}", label, bold=True)
        for col in ("E", "F", "G"):
            rng = f"{col}{first}:{col}{last}"
            f = (f"=MEDIAN({rng})" if k == 1
                 else f"=QUARTILE({rng},{1 if k == 0 else 3})")
            set_cell(ws, f"{col}{stats + k}", f, None, FMT_X,
                     border=THIN_TOP if k == 0 else None)

    imp = stats + 4
    set_cell(ws, f"A{imp}", "Implied value per share "
             "(target LTM metrics × peer multiples)", bold=True)
    hdr = imp + 1
    for j, h in enumerate(("Method", "25th pct", "Median", "75th pct")):
        set_cell(ws, f"{col_letter(j, 1)}{hdr}", h, HDR_FONT, fill=HDR_FILL)
    q1, med, q3 = stats, stats + 1, stats + 2
    methods = [
        ("EV/EBITDA", "E", f"IS!${last_hist}${IS['ebitda']}"),
        ("EV/Revenue", "F", f"IS!${last_hist}${IS['rev']}"),
    ]
    r = hdr + 1
    for label, col, metric in methods:
        set_cell(ws, f"A{r}", label)
        for j, srow in enumerate((q1, med, q3)):
            set_cell(ws, f"{col_letter(j + 1, 1)}{r}",
                     f"=({col}{srow}*{metric}-DCF!$B${D['net_debt']})"
                     f"/DCF!$B${D['shares']}", None, FMT_PS)
        r += 1
    set_cell(ws, f"A{r}", "P/E")
    for j, srow in enumerate((q1, med, q3)):
        set_cell(ws, f"{col_letter(j + 1, 1)}{r}",
                 f"=G{srow}*IS!${last_hist}${IS['ni']}/DCF!$B${D['shares']}",
                 None, FMT_PS)
    set_widths(ws, {"A": 14, "B": 26, "C": 14, "D": 14, "E": 12, "F": 12,
                    "G": 12})
    return {"imp_first": hdr + 1, "imp_last": r}


def build_summary_sheet(ws, data, comps_rows):
    sheet_title(ws, f"{data['ticker']} — Valuation Summary",
                "Football field: implied value per share by methodology ($).")
    for j, h in enumerate(("Method", "Low", "High", "Range (chart)")):
        set_cell(ws, f"{col_letter(j, 1)}4", h, HDR_FONT, fill=HDR_FILL)

    imp = comps_rows["imp_first"]
    sf, sl = D["sens_first"], D["sens_first"] + 4
    methods = [
        ("52-week trading range", data["wk52_low"], data["wk52_high"], BLUE),
        ("Comps: EV/EBITDA (25th–75th)",
         f"=Comps!B{imp}", f"=Comps!D{imp}", None),
        ("Comps: EV/Revenue (25th–75th)",
         f"=Comps!B{imp + 1}", f"=Comps!D{imp + 1}", None),
        ("Comps: P/E (25th–75th)",
         f"=Comps!B{imp + 2}", f"=Comps!D{imp + 2}", None),
        ("DCF: sensitivity min–max",
         f"=MIN(DCF!C{sf}:G{sl})", f"=MAX(DCF!C{sf}:G{sl})", None),
        ("DCF: Gordon base case",
         f"=DCF!B{D['ps']}*0.999", f"=DCF!B{D['ps']}*1.001", None),
        ("DCF: exit-multiple base case",
         f"=DCF!B{D['ps_exit']}*0.999", f"=DCF!B{D['ps_exit']}*1.001", None),
        ("Current price",
         f"=DCF!B{D['cur']}*0.999", f"=DCF!B{D['cur']}*1.001", None),
    ]
    for i, (label, lo, hi, font) in enumerate(methods, start=5):
        set_cell(ws, f"A{i}", label)
        set_cell(ws, f"B{i}", lo, font, FMT_PS)
        set_cell(ws, f"C{i}", hi, font, FMT_PS)
        set_cell(ws, f"D{i}", f"=C{i}-B{i}", None, FMT_PS)

    last = 4 + len(methods)
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Football Field — Implied Value per Share ($)"
    chart.height, chart.width = 9, 20
    base = Reference(ws, min_col=2, min_row=4, max_row=last)
    rng = Reference(ws, min_col=4, min_row=4, max_row=last)
    chart.add_data(base, titles_from_data=True)
    chart.add_data(rng, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=5, max_row=last))
    chart.series[0].graphicalProperties.noFill = True
    chart.series[1].graphicalProperties.solidFill = NAVY
    chart.legend = None
    ws.add_chart(chart, "F4")
    set_widths(ws, {"A": 32, "B": 12, "C": 12, "D": 14})


def build_dcf_model(data, assumptions, out_path, peers=None, wacc_seeds=None):
    """Three-statement workbook + DCF (+ Comps + Summary when peers given)."""
    from openpyxl import load_workbook

    if not data.get("price") or not data.get("shares_mm"):
        raise ValueError(f"{data['ticker']}: no market data (price/shares) — "
                         "cannot build a DCF. Check the FMP profile payload.")
    build_three_statement_wb(data, assumptions, out_path)
    wb = load_workbook(out_path)  # reopen to append valuation sheets

    w = wacc_seeds or derive_wacc_seeds(data, assumptions, peers)
    n_hist = data["n_hist"]
    build_dcf_sheet(wb.create_sheet("DCF"), data, w, n_hist)
    if peers:
        comps_rows = build_comps_sheet(wb.create_sheet("Comps"), data, peers,
                                       n_hist)
        build_summary_sheet(wb.create_sheet("Summary"), data, comps_rows)
    wb.save(out_path)
    return out_path
