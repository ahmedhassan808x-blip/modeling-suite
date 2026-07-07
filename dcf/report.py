"""
Presentation & export layer for the DCF workbook.

    python3 -m dcf.report AAPL_dcf.xlsx

Produces {stem}_deck.pptx, {stem}_memo.docx, {stem}_deck.pdf next to the
workbook. Numbers come from the recalculated workbook; the scenario slide
re-runs the whole model per case via the in-workbook toggle.
"""

import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from dcf.model_builder import D, SENS_COLS  # noqa: E402
from shared import charts  # noqa: E402
from shared.exports.docx_memo import Memo  # noqa: E402
from shared.exports.pdf import to_pdf  # noqa: E402
from shared.exports.pptx_deck import Deck  # noqa: E402
from shared.recalc import recalculate  # noqa: E402
from three_statement.report import (  # noqa: E402
    FC, N_HIST, _driver_table, _fmt, extract as extract_3s, scenario_runs,
)

DCF_LIMITS = [
    "A 5-year explicit DCF is structurally conservative on premium "
    "compounders — most of the value sits in the terminal assumption, and "
    "the sensitivity grid shows exactly how much. Treat 'expensive' "
    "readings on quality mega-caps as a property of the method before "
    "calling them a signal.",
    "Risk-free rate and equity risk premium are labeled inputs, not live "
    "feeds — update the 10Y before using the output.",
    "No mid-year convention or stub periods; single currency; peer set is "
    "a judgment the analyst owns.",
    "All three-statement limitations flow through (no buybacks, "
    "beginning-balance interest, flat 'other' lines).",
]


def _dcf_probes():
    p = [f"DCF!B{r}" for r in set(D.values()) if r < D["sens_hdr"]]
    p += [f"DCF!{c}{D['sens_hdr']}" for c in SENS_COLS]
    for i in range(5):
        r = D["sens_first"] + i
        p += [f"DCF!B{r}"] + [f"DCF!{c}{r}" for c in SENS_COLS]
    p += [f"DCF!{c}{D['ufcf']}" for c in FC]
    return p


def extract_dcf(xlsx) -> dict:
    base = extract_3s(xlsx)
    wb = load_workbook(xlsx)
    has_summary = "Summary" in wb.sheetnames
    probes = _dcf_probes()
    if has_summary:
        probes += [f"Summary!{col}{r}" for r in range(5, 13)
                   for col in ("B", "C")]
        field_labels = [wb["Summary"][f"A{r}"].value for r in range(5, 13)]
    res = recalculate(xlsx, probe_cells=probes)
    if not res.ok:
        raise RuntimeError(f"{xlsx}: formula errors — refusing to report.\n"
                           + res.summary())
    v = res.values

    def b(key):
        return v[f"DCF!B{D[key]}"]

    sens = dict(
        g=[v[f"DCF!{c}{D['sens_hdr']}"] for c in SENS_COLS],
        wacc=[v[f"DCF!B{D['sens_first'] + i}"] for i in range(5)],
        grid=[[v[f"DCF!{c}{D['sens_first'] + i}"] for c in SENS_COLS]
              for i in range(5)],
    )
    if has_summary:
        field = [(lbl, v[f"Summary!B{r}"], v[f"Summary!C{r}"])
                 for lbl, r in zip(field_labels, range(5, 13))
                 if lbl and "Current price" not in lbl]
    else:
        lo = min(min(r) for r in sens["grid"])
        hi = max(max(r) for r in sens["grid"])
        field = [("DCF sensitivity min–max", lo, hi),
                 ("DCF Gordon base", b("ps") * 0.995, b("ps") * 1.005),
                 ("DCF exit-multiple base", b("ps_exit") * 0.995,
                  b("ps_exit") * 1.005)]
    base.update(dcf={k: b(k) for k in ("wacc", "ke", "rf", "erp", "beta",
                                       "kd", "dw", "tg", "sum_pv", "tv",
                                       "pv_tv", "ev", "net_debt", "eq", "ps",
                                       "cur", "upside", "exit_mult",
                                       "ps_exit")},
                sens=sens, field=field,
                ufcf=[v[f"DCF!{c}{D['ufcf']}"] for c in FC])
    return base


def _sens_table(sens):
    rows = [["WACC \\ g"] + [f"{g:.2%}" for g in sens["g"]]]
    for w, grid_row in zip(sens["wacc"], sens["grid"]):
        rows.append([f"{w:.2%}"] + [f"${x:,.0f}" for x in grid_row])
    return rows


def _wacc_table(dc):
    return [["WACC build-up", ""],
            ["Risk-free rate", f"{dc['rf']:.2%}"],
            ["Beta", f"{dc['beta']:.2f}"],
            ["Equity risk premium", f"{dc['erp']:.2%}"],
            ["Cost of equity", f"{dc['ke']:.2%}"],
            ["Pre-tax cost of debt", f"{dc['kd']:.2%}"],
            ["Debt / (D+E)", f"{dc['dw']:.1%}"],
            ["WACC", f"{dc['wacc']:.2%}"],
            ["Terminal growth (g)", f"{dc['tg']:.2%}"]]


def build_reports(xlsx, outdir=None):
    xlsx = Path(xlsx)
    outdir = Path(outdir) if outdir else xlsx.parent
    d = extract_dcf(xlsx)
    dc = d["dcf"]
    scen = scenario_runs(xlsx, extra_probes=[f"DCF!B{D['ps']}"])
    years_fc = d["years"][N_HIST:]

    tmp = Path(tempfile.mkdtemp())
    png_field = charts.football_field(d["field"], dc["cur"], tmp / "ff.png")
    png_ufcf = charts.ufcf_bars(years_fc, d["ufcf"], tmp / "ufcf.png")
    png_tv = charts.tv_split(dc["sum_pv"], dc["pv_tv"], tmp / "tv.png")

    tv_share = dc["pv_tv"] / (dc["sum_pv"] + dc["pv_tv"])
    summary_bullets = [
        f"Gordon-growth DCF implies {_fmt(dc['ps'], 'ps')}/share "
        f"({dc['upside']:+.1%} vs. market {_fmt(dc['cur'], 'ps')}).",
        f"Exit-multiple cross-check at {dc['exit_mult']:.1f}x EV/EBITDA "
        f"implies {_fmt(dc['ps_exit'], 'ps')}/share.",
        f"WACC {dc['wacc']:.2%} (cost of equity {dc['ke']:.2%}, "
        f"{dc['dw']:.0%} debt weight); terminal growth {dc['tg']:.2%}.",
        f"{tv_share:.0%} of enterprise value sits in the terminal value — "
        "the honest measure of how much this DCF depends on assumptions "
        "nobody can know.",
        "UFCF is linked live to a balanced three-statement forecast — not a "
        "standalone margin build.",
    ]
    scen_rows = [["Scenario", "Implied $/share", "vs. market"]]
    for lbl in ("Bear", "Base", "Bull"):
        ps = scen[lbl][f"DCF!B{D['ps']}"]
        scen_rows.append([lbl, f"${ps:,.2f}", f"{ps / dc['cur'] - 1:+.1%}"])

    footer = (f"{d['ticker']} — DCF valuation | modeling-suite | "
              "$mm unless noted | AI-assembled from the recalculated model — "
              "verify assumptions before use")

    deck = Deck(footer)
    deck.title_slide(f"{d['name']} — DCF Valuation",
                     f"FCFF discounted cash flow, linked three-statement "
                     f"model | {d['ticker']}",
                     "Live Excel model: blue = input, black = formula, "
                     "green = cross-sheet link")
    s = deck.content_slide("Executive summary — valuation")
    deck.add_bullets(s, summary_bullets, 0.5, 1.3, 5.2, 5.4, size=12)
    deck.add_image(s, png_field, 5.9, 1.4, 7.0)

    s = deck.content_slide("Unlevered free cash flow & value composition")
    deck.add_image(s, png_ufcf, 0.5, 1.4, 6.4)
    deck.add_image(s, png_tv, 7.2, 1.8, 5.6)

    s = deck.content_slide("Key assumptions")
    deck.add_table(s, _wacc_table(dc), 0.5, 1.4, 4.4, first_col_w=2.4)
    deck.add_table(s, _driver_table(d, years_fc), 5.4, 1.4, 7.5,
                   first_col_w=2.4)

    s = deck.content_slide("Sensitivities & scenarios")
    deck.add_table(s, _sens_table(d["sens"]), 0.5, 1.5, 6.2, first_col_w=1.4)
    deck.add_table(s, scen_rows, 7.2, 1.5, 5.4)
    deck.add_bullets(s, [
        "Scenario values re-run the entire linked model (statements, "
        "revolver, DCF) via the in-workbook toggle.",
    ], 7.2, 3.4, 5.4, 1.6, size=10)

    s = deck.content_slide("Risks & limitations")
    deck.add_bullets(s, DCF_LIMITS, 0.5, 1.4, 12.0, 5.2, size=12)
    pptx_path = deck.save(outdir / f"{xlsx.stem}_deck.pptx")

    memo = Memo(f"{d['name']} ({d['ticker']}) — DCF Valuation Memo",
                "FCFF DCF on a linked three-statement forecast | "
                "modeling-suite")
    memo.heading("Executive summary")
    memo.bullets(summary_bullets)
    memo.image(png_field)
    memo.heading("Free cash flow & value composition")
    memo.image(png_ufcf)
    memo.image(png_tv)
    memo.heading("Assumptions")
    memo.table(_wacc_table(dc))
    memo.table(_driver_table(d, years_fc))
    memo.heading("Sensitivities & scenarios")
    memo.table(_sens_table(d["sens"]))
    memo.table(scen_rows)
    memo.heading("Risks & limitations")
    memo.bullets(DCF_LIMITS)
    memo.para("Generated by modeling-suite from the recalculated workbook — "
              "figures cannot diverge from the model.", small=True)
    docx_path = memo.save(outdir / f"{xlsx.stem}_memo.docx")

    pdf_path = to_pdf(pptx_path, outdir)
    shutil.rmtree(tmp, ignore_errors=True)
    return dict(pptx=pptx_path, docx=docx_path, pdf=pdf_path)


def main():
    import argparse
    ap = argparse.ArgumentParser(description="Build deck/memo/PDF from a DCF "
                                 "workbook.")
    ap.add_argument("workbook")
    ap.add_argument("--outdir")
    args = ap.parse_args()
    for k, p in build_reports(args.workbook, args.outdir).items():
        print(f"{k}: {p}")


if __name__ == "__main__":
    main()
