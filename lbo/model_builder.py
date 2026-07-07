"""
LBO workbook builder — extends the three-statement model.

The operating engine IS the linked three-statement forecast: EBITDA, D&A,
capex and ΔNWC flow into the LBO FCF build as green links, so the scenario
toggle on Assumptions reprices the IRR. The LBO layers on top:

  Deal       entry multiple (seeded from the company's actual market
             EV/EBITDA), sources & uses, tranche sizing and pricing
  LBO        FCF-for-debt-service build; revolver (draws to cover any
             shortfall vs. mandatory amort, swept down first), senior
             (mandatory amort + cash sweep), subordinated (sweep last)
  Returns    exit EV -> equity, IRR (real IRR() over the flows) and MOIC,
             value-creation bridge, exit-multiple x exit-year sensitivity
  ChecksLBO  sources = uses, non-negative balances everywhere, and
             IRR() == closed-form MOIC^(1/n)-1 cross-check

The revolver exists because a fixed mandatory amort can exceed FCF in weak
years (the bear scenario does exactly this) — a real deal funds that gap
with a revolver, and so does this model. Revolver usage is the visible
feasibility signal; the drawn balance costs interest and reduces exit
equity through net debt.

Interest on beginning-of-period balances (suite convention: no circular
references, so the recalc gate can verify every cell).

Honesty note on sensitivities: leverage cannot be a LIVE grid axis — each
leverage point implies a different sweep path, which no single cell formula
can honestly compute. The in-workbook grid is exit multiple x exit year
(fully live); the report layer adds leverage x exit multiple by genuinely
re-running the model per leverage point.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from shared.excel_utils import (  # noqa: E402
    BLUE, BOLD, GRAY, HDR_FILL, HDR_FONT, INPUT_HL, THIN_TOP,
    FMT_PCT, FMT_USD, FMT_X, col_letter, set_cell, set_widths, sheet_title,
)
from three_statement.model_builder import (  # noqa: E402
    ASM, CF, IS, N_FC, WIDTHS, build_model as build_three_statement_wb,
)

# Row maps
DEAL = dict(entry_mult=5, ltm_ebitda=6, entry_ev=7, fees_pct=8, fees=9,
            senior_x=11, sub_x=12, senior_amt=13, sub_amt=14,
            senior_rate=15, sub_rate=16, rv_rate=17, amort_pct=18,
            exit_mult=19,
            u_ev=23, u_fees=24, u_tot=25,
            s_senior=28, s_sub=29, s_equity=30, s_tot=31)
L = dict(ebitda=5, da=6, ebit=7, int_rv=8, int_sr=9, int_sub=10, ebt=11,
         tax=12, ni=13, da_back=14, capex=15, nwc=16, fcf=17,
         rv_beg=20, rv_draw=21, rv_pay=22, rv_end=23,
         sr_beg=26, sr_mand=27, sr_sweep=28, sr_end=29,
         sub_beg=32, sub_sweep=33, sub_end=34,
         cash_beg=37, cash_chg=38, cash_end=39,
         debt_end=41, net_debt=42)
R = dict(exit_ebitda=5, exit_mult=6, exit_ev=7, nd_exit=8, exit_eq=9,
         entry_eq=10, moic=11, irr=12, flows=14,
         br_entry=17, br_growth=18, br_mult=19, br_delev=20, br_fees=21,
         br_exit=22, sens_hdr=25, sens_first=26)
SENS_YEARS = (3, 4, 5)
CHECKS_PASS_CELL = "ChecksLBO!B11"


def _fc_cols(n_hist):
    return [col_letter(n_hist + i) for i in range(N_FC)]


def build_deal(ws, data, lb, n_hist):
    last_hist = col_letter(n_hist - 1)
    sheet_title(ws, f"{data['ticker']} — LBO Transaction & Sources / Uses",
                "Entry seeded from the company's actual market EV/EBITDA; "
                "exit seeded flat to entry. All figures $mm.")
    set_cell(ws, "A4", "Transaction inputs", None, bold=True)
    inputs = [
        ("entry_mult", "Entry EV / LTM EBITDA (seed: market implied)",
         lb.entry_mult, FMT_X, True),
        ("ltm_ebitda", "LTM EBITDA (last actual)",
         f"=IS!{last_hist}{IS['ebitda']}", FMT_USD, False),
        ("entry_ev", "Entry enterprise value",
         f"=B{DEAL['entry_mult']}*B{DEAL['ltm_ebitda']}", FMT_USD, False),
        ("fees_pct", "Transaction fees (% of EV)", lb.fees_pct, FMT_PCT, True),
        ("fees", "Transaction fees",
         f"=B{DEAL['entry_ev']}*B{DEAL['fees_pct']}", FMT_USD, False),
        ("senior_x", "Senior debt (x LTM EBITDA)", lb.senior_x, FMT_X, True),
        ("sub_x", "Subordinated debt (x LTM EBITDA)", lb.sub_x, FMT_X, True),
        ("senior_amt", "Senior debt",
         f"=B{DEAL['senior_x']}*B{DEAL['ltm_ebitda']}", FMT_USD, False),
        ("sub_amt", "Subordinated debt",
         f"=B{DEAL['sub_x']}*B{DEAL['ltm_ebitda']}", FMT_USD, False),
        ("senior_rate", "Senior interest rate", lb.senior_rate, FMT_PCT, True),
        ("sub_rate", "Subordinated interest rate", lb.sub_rate, FMT_PCT, True),
        ("rv_rate", "Revolver interest rate", lb.rv_rate, FMT_PCT, True),
        ("amort_pct", "Senior mandatory amort (%/yr of original)",
         lb.amort_pct, FMT_PCT, True),
        ("exit_mult", "Exit EV / EBITDA (seed: flat to entry)",
         lb.exit_mult, FMT_X, True),
    ]
    for key, label, val, fmt, is_input in inputs:
        set_cell(ws, f"A{DEAL[key]}", label)
        c = set_cell(ws, f"B{DEAL[key]}", val, BLUE if is_input else None, fmt)
        if is_input:
            c.fill = INPUT_HL

    set_cell(ws, f"A{DEAL['u_ev'] - 1}", "Uses", HDR_FONT, fill=HDR_FILL)
    set_cell(ws, f"B{DEAL['u_ev'] - 1}", "$mm", HDR_FONT, fill=HDR_FILL)
    set_cell(ws, f"A{DEAL['u_ev']}",
             "Purchase enterprise value (equity + refinance net debt)")
    set_cell(ws, f"B{DEAL['u_ev']}", f"=B{DEAL['entry_ev']}", None, FMT_USD)
    set_cell(ws, f"A{DEAL['u_fees']}", "Transaction fees")
    set_cell(ws, f"B{DEAL['u_fees']}", f"=B{DEAL['fees']}", None, FMT_USD)
    set_cell(ws, f"A{DEAL['u_tot']}", "Total uses", None, bold=True)
    set_cell(ws, f"B{DEAL['u_tot']}",
             f"=B{DEAL['u_ev']}+B{DEAL['u_fees']}", None, FMT_USD, bold=True,
             border=THIN_TOP)

    set_cell(ws, f"A{DEAL['s_senior'] - 1}", "Sources", HDR_FONT,
             fill=HDR_FILL)
    set_cell(ws, f"B{DEAL['s_senior'] - 1}", "$mm", HDR_FONT, fill=HDR_FILL)
    set_cell(ws, f"A{DEAL['s_senior']}", "Senior debt")
    set_cell(ws, f"B{DEAL['s_senior']}", f"=B{DEAL['senior_amt']}", None,
             FMT_USD)
    set_cell(ws, f"A{DEAL['s_sub']}", "Subordinated debt")
    set_cell(ws, f"B{DEAL['s_sub']}", f"=B{DEAL['sub_amt']}", None, FMT_USD)
    set_cell(ws, f"A{DEAL['s_equity']}", "Sponsor equity (plug)", None,
             bold=True)
    set_cell(ws, f"B{DEAL['s_equity']}",
             f"=B{DEAL['u_tot']}-B{DEAL['s_senior']}-B{DEAL['s_sub']}",
             None, FMT_USD, bold=True)
    set_cell(ws, f"A{DEAL['s_tot']}", "Total sources", None, bold=True)
    set_cell(ws, f"B{DEAL['s_tot']}",
             f"=SUM(B{DEAL['s_senior']}:B{DEAL['s_equity']})", None, FMT_USD,
             bold=True, border=THIN_TOP)
    set_widths(ws, dict(WIDTHS, A=48))


def build_lbo(ws, data, n_hist):
    fc = _fc_cols(n_hist)
    sheet_title(ws, f"{data['ticker']} — LBO Cash Flow & Debt Schedule ($mm)",
                "Operating lines are green links into the three-statement "
                "forecast (scenario toggle applies). Interest on beginning "
                "balances. Sweep order: revolver, senior, subordinated; the "
                "revolver funds any shortfall vs. mandatory amort.")
    for t, c in enumerate(fc, start=1):
        set_cell(ws, f"{c}4", f"Year {t}", BOLD)

    labels = [
        ("ebitda", "EBITDA", False), ("da", "D&A", False),
        ("ebit", "EBIT", True), ("int_rv", "Interest — revolver", False),
        ("int_sr", "Interest — senior", False),
        ("int_sub", "Interest — subordinated", False),
        ("ebt", "Pre-tax income", False),
        ("tax", "Cash taxes (no benefit on losses)", False),
        ("ni", "Net income", True), ("da_back", "(+) D&A", False),
        ("capex", "(–) Capex", False), ("nwc", "(–) Increase in NWC", False),
        ("fcf", "FCF available for debt service", True),
        ("rv_beg", "Revolver — beginning", False),
        ("rv_draw", "Revolver — draw (funds shortfall)", False),
        ("rv_pay", "Revolver — (paydown)", False),
        ("rv_end", "Revolver — ending", True),
        ("sr_beg", "Senior — beginning", False),
        ("sr_mand", "Senior — mandatory amort", False),
        ("sr_sweep", "Senior — cash sweep", False),
        ("sr_end", "Senior — ending", True),
        ("sub_beg", "Subordinated — beginning", False),
        ("sub_sweep", "Subordinated — cash sweep", False),
        ("sub_end", "Subordinated — ending", True),
        ("cash_beg", "Cash — beginning", False),
        ("cash_chg", "Cash build", False),
        ("cash_end", "Cash — ending", True),
        ("debt_end", "Total debt — ending", True),
        ("net_debt", "Net debt — ending", True),
    ]
    for key, label, bold in labels:
        set_cell(ws, f"A{L[key]}", label, None, bold=bold)
    for row, txt in ((19, "Revolver"), (25, "Senior tranche"),
                     (31, "Subordinated tranche"), (36, "Cash")):
        set_cell(ws, f"A{row}", txt, None, bold=True)

    for j, c in enumerate(fc):
        p = col_letter(n_hist + j - 1)
        set_cell(ws, f"{c}{L['ebitda']}", f"=IS!{c}{IS['ebitda']}", None,
                 FMT_USD)
        set_cell(ws, f"{c}{L['da']}", f"=IS!{c}{IS['da']}", None, FMT_USD)
        set_cell(ws, f"{c}{L['ebit']}", f"={c}{L['ebitda']}+{c}{L['da']}",
                 None, FMT_USD, bold=True)
        set_cell(ws, f"{c}{L['int_rv']}",
                 f"=-Deal!$B${DEAL['rv_rate']}*{c}{L['rv_beg']}", None,
                 FMT_USD)
        set_cell(ws, f"{c}{L['int_sr']}",
                 f"=-Deal!$B${DEAL['senior_rate']}*{c}{L['sr_beg']}", None,
                 FMT_USD)
        set_cell(ws, f"{c}{L['int_sub']}",
                 f"=-Deal!$B${DEAL['sub_rate']}*{c}{L['sub_beg']}", None,
                 FMT_USD)
        set_cell(ws, f"{c}{L['ebt']}",
                 f"=SUM({c}{L['ebit']}:{c}{L['int_sub']})", None, FMT_USD)
        set_cell(ws, f"{c}{L['tax']}",
                 f"=-MAX(0,{c}{L['ebt']})*Assumptions!{c}{ASM['tax']}", None,
                 FMT_USD)
        set_cell(ws, f"{c}{L['ni']}", f"={c}{L['ebt']}+{c}{L['tax']}", None,
                 FMT_USD, bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{L['da_back']}", f"=-{c}{L['da']}", None, FMT_USD)
        set_cell(ws, f"{c}{L['capex']}", f"=CF!{c}{CF['capex']}", None,
                 FMT_USD)
        nwc = "+".join(f"CF!{c}{CF[k]}"
                       for k in ("d_ar", "d_inv", "d_oca", "d_ap", "d_ocl"))
        set_cell(ws, f"{c}{L['nwc']}", f"={nwc}", None, FMT_USD)
        set_cell(ws, f"{c}{L['fcf']}",
                 f"=SUM({c}{L['ni']}:{c}{L['nwc']})", None, FMT_USD,
                 bold=True, border=THIN_TOP)

        # available after mandatory amort = FCF + sr_mand (mand is negative)
        avail = f"{c}{L['fcf']}+{c}{L['sr_mand']}"
        if j == 0:
            set_cell(ws, f"{c}{L['rv_beg']}", 0, BLUE, FMT_USD)
        else:
            set_cell(ws, f"{c}{L['rv_beg']}", f"={p}{L['rv_end']}", None,
                     FMT_USD)
        set_cell(ws, f"{c}{L['rv_draw']}", f"=MAX(0,-({avail}))", None,
                 FMT_USD)
        set_cell(ws, f"{c}{L['rv_pay']}",
                 f"=-MIN({c}{L['rv_beg']},MAX(0,{avail}))", None, FMT_USD)
        set_cell(ws, f"{c}{L['rv_end']}",
                 f"={c}{L['rv_beg']}+{c}{L['rv_draw']}+{c}{L['rv_pay']}",
                 None, FMT_USD, bold=True, border=THIN_TOP)

        sr_beg = (f"=Deal!$B${DEAL['senior_amt']}" if j == 0
                  else f"={p}{L['sr_end']}")
        set_cell(ws, f"{c}{L['sr_beg']}", sr_beg, None, FMT_USD)
        set_cell(ws, f"{c}{L['sr_mand']}",
                 f"=-MIN({c}{L['sr_beg']},Deal!$B${DEAL['amort_pct']}"
                 f"*Deal!$B${DEAL['senior_amt']})", None, FMT_USD)
        set_cell(ws, f"{c}{L['sr_sweep']}",
                 f"=-MIN({c}{L['sr_beg']}+{c}{L['sr_mand']},"
                 f"MAX(0,{avail}+{c}{L['rv_pay']}))", None, FMT_USD)
        set_cell(ws, f"{c}{L['sr_end']}",
                 f"={c}{L['sr_beg']}+{c}{L['sr_mand']}+{c}{L['sr_sweep']}",
                 None, FMT_USD, bold=True, border=THIN_TOP)

        sub_beg = (f"=Deal!$B${DEAL['sub_amt']}" if j == 0
                   else f"={p}{L['sub_end']}")
        set_cell(ws, f"{c}{L['sub_beg']}", sub_beg, None, FMT_USD)
        set_cell(ws, f"{c}{L['sub_sweep']}",
                 f"=-MIN({c}{L['sub_beg']},MAX(0,{avail}"
                 f"+{c}{L['rv_pay']}+{c}{L['sr_sweep']}))", None, FMT_USD)
        set_cell(ws, f"{c}{L['sub_end']}",
                 f"={c}{L['sub_beg']}+{c}{L['sub_sweep']}", None, FMT_USD,
                 bold=True, border=THIN_TOP)

        if j == 0:
            set_cell(ws, f"{c}{L['cash_beg']}", 0, BLUE, FMT_USD)
        else:
            set_cell(ws, f"{c}{L['cash_beg']}", f"={p}{L['cash_end']}", None,
                     FMT_USD)
        set_cell(ws, f"{c}{L['cash_chg']}",
                 f"={avail}+{c}{L['rv_draw']}+{c}{L['rv_pay']}"
                 f"+{c}{L['sr_sweep']}+{c}{L['sub_sweep']}", None, FMT_USD)
        set_cell(ws, f"{c}{L['cash_end']}",
                 f"={c}{L['cash_beg']}+{c}{L['cash_chg']}", None, FMT_USD,
                 bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{L['debt_end']}",
                 f"={c}{L['rv_end']}+{c}{L['sr_end']}+{c}{L['sub_end']}",
                 None, FMT_USD, bold=True)
        set_cell(ws, f"{c}{L['net_debt']}",
                 f"={c}{L['debt_end']}-{c}{L['cash_end']}", None, FMT_USD,
                 bold=True)
    set_widths(ws, dict(WIDTHS, A=38))


def build_returns(ws, data, n_hist):
    fc = _fc_cols(n_hist)
    last = fc[-1]
    sheet_title(ws, f"{data['ticker']} — LBO Returns",
                "Exit at end of Year 5. IRR from actual flows; bridge "
                "decomposes value creation. Sensitivity is exit multiple × "
                "exit year (live); leverage × exit lives in the report, "
                "computed by re-running the model.")
    rows = [
        ("exit_ebitda", "Exit-year EBITDA", f"=IS!{last}{IS['ebitda']}",
         FMT_USD),
        ("exit_mult", "Exit multiple", f"=Deal!B{DEAL['exit_mult']}", FMT_X),
        ("exit_ev", "Exit enterprise value",
         f"=B{R['exit_ebitda']}*B{R['exit_mult']}", FMT_USD),
        ("nd_exit", "Net debt at exit", f"=LBO!{last}{L['net_debt']}",
         FMT_USD),
        ("exit_eq", "Equity value at exit",
         f"=B{R['exit_ev']}-B{R['nd_exit']}", FMT_USD),
        ("entry_eq", "Sponsor equity at entry",
         f"=Deal!B{DEAL['s_equity']}", FMT_USD),
        ("moic", "MOIC", f"=B{R['exit_eq']}/B{R['entry_eq']}", FMT_X),
        ("irr", "IRR (5-year)",
         f"=IRR(C{R['flows']}:H{R['flows']})", FMT_PCT),
    ]
    for key, label, val, fmt in rows:
        set_cell(ws, f"A{R[key]}", label, None,
                 bold=key in ("moic", "irr", "exit_eq"))
        set_cell(ws, f"B{R[key]}", val, None, fmt,
                 bold=key in ("moic", "irr"))

    set_cell(ws, f"A{R['flows']}", "Sponsor cash flows (entry → exit)", GRAY)
    set_cell(ws, f"C{R['flows']}", f"=-B{R['entry_eq']}", None, FMT_USD)
    for i in range(4):
        set_cell(ws, f"{chr(ord('D') + i)}{R['flows']}", 0, BLUE, FMT_USD)
    set_cell(ws, f"H{R['flows']}", f"=B{R['exit_eq']}", None, FMT_USD)

    set_cell(ws, f"A{R['br_entry'] - 1}", "Value-creation bridge ($mm)", None,
             bold=True)
    bridge = [
        ("br_entry", "Sponsor equity at entry", f"=B{R['entry_eq']}"),
        ("br_growth", "(+) EBITDA growth (at entry multiple)",
         f"=(B{R['exit_ebitda']}-Deal!B{DEAL['ltm_ebitda']})"
         f"*Deal!B{DEAL['entry_mult']}"),
        ("br_mult", "(+) Multiple expansion / (contraction)",
         f"=(B{R['exit_mult']}-Deal!B{DEAL['entry_mult']})"
         f"*B{R['exit_ebitda']}"),
        ("br_delev", "(+) Deleveraging & cash build",
         f"=Deal!B{DEAL['senior_amt']}+Deal!B{DEAL['sub_amt']}"
         f"-B{R['nd_exit']}"),
        ("br_fees", "(–) Fees & other",
         f"=B{R['br_exit']}-B{R['br_entry']}-B{R['br_growth']}"
         f"-B{R['br_mult']}-B{R['br_delev']}"),
        ("br_exit", "Equity value at exit", f"=B{R['exit_eq']}"),
    ]
    for key, label, val in bridge:
        set_cell(ws, f"A{R[key]}", label)
        set_cell(ws, f"B{R[key]}", val, None, FMT_USD,
                 bold=key in ("br_entry", "br_exit"))

    set_cell(ws, f"A{R['sens_hdr'] - 1}",
             "Sensitivity — IRR by exit multiple × exit year", None, bold=True)
    set_cell(ws, f"B{R['sens_hdr']}", "Exit x \\ year", BOLD)
    for k, t in enumerate(SENS_YEARS):
        set_cell(ws, f"{chr(ord('C') + k)}{R['sens_hdr']}", f"Year {t}", BOLD)
    for i in range(5):
        r = R["sens_first"] + i
        set_cell(ws, f"B{r}", f"=Deal!$B${DEAL['exit_mult']}+{i - 2}*1",
                 None, FMT_X)
        for k, t in enumerate(SENS_YEARS):
            yc = col_letter(n_hist + t - 1)
            sc = chr(ord("C") + k)
            eq = (f"($B{r}*IS!{yc}${IS['ebitda']}"
                  f"-LBO!{yc}${L['net_debt']})")
            set_cell(ws, f"{sc}{r}",
                     f"=IF({eq}<=0,-1,({eq}/$B${R['entry_eq']})^(1/{t})-1)",
                     None, FMT_PCT)
    set_widths(ws, dict(WIDTHS, A=42))


def build_checks_lbo(ws, data, n_hist):
    fc = _fc_cols(n_hist)
    sheet_title(ws, f"{data['ticker']} — LBO Integrity Checks",
                "Every value zero and B11 = PASS; asserted by the automated "
                "recalc gate.")
    set_cell(ws, "A4", "Sources − uses")
    set_cell(ws, "B4", f"=Deal!B{DEAL['s_tot']}-Deal!B{DEAL['u_tot']}", None,
             "0.000")
    set_cell(ws, "A5", "IRR vs closed-form MOIC^(1/5)−1")
    set_cell(ws, "B5",
             f"=Returns!B{R['irr']}-(Returns!B{R['moic']}^(1/5)-1)", None,
             "0.0000")
    checks = [("Revolver ≥ 0", L["rv_end"], 6),
              ("Senior ≥ 0", L["sr_end"], 7),
              ("Subordinated ≥ 0", L["sub_end"], 8),
              ("Cash ≥ 0", L["cash_end"], 9)]
    for label, lrow, row in checks:
        set_cell(ws, f"A{row}", label)
        for c in fc:
            set_cell(ws, f"{c}{row}",
                     f"=IF(LBO!{c}{lrow}>=-0.001,0,1)", None, "0")
    set_cell(ws, "A11", "ALL LBO CHECKS", None, bold=True)
    cell = set_cell(
        ws, "B11",
        f'=IF(ABS(B4)+ABS(B5)+SUM({fc[0]}6:{fc[-1]}9)<0.005,"PASS","FAIL")')
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    set_widths(ws, dict(WIDTHS, A=42))


def build_lbo_model(data, assumptions, lbo_assumptions, out_path):
    """Three-statement workbook + Deal / LBO / Returns / ChecksLBO."""
    build_three_statement_wb(data, assumptions, out_path)
    wb = load_workbook(out_path)
    n_hist = data["n_hist"]
    build_deal(wb.create_sheet("Deal"), data, lbo_assumptions, n_hist)
    build_lbo(wb.create_sheet("LBO"), data, n_hist)
    build_returns(wb.create_sheet("Returns"), data, n_hist)
    build_checks_lbo(wb.create_sheet("ChecksLBO"), data, n_hist)
    wb.save(out_path)
    return out_path
