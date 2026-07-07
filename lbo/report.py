"""
Presentation & export layer for the LBO workbook.

    python3 -m lbo.report AAPL_lbo.xlsx

Produces {stem}_deck.pptx, {stem}_memo.docx, {stem}_deck.pdf. Numbers come
from the recalculated workbook. The leverage x exit-multiple IRR table is
computed by GENUINELY re-running the model per leverage point (editing the
Deal inputs and recalculating) — not by an approximation formula, because
each leverage level implies a different sweep path.
"""

import re
import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from lbo.model_builder import CHECKS_PASS_CELL, DEAL, L, R  # noqa: E402
from shared import charts  # noqa: E402
from shared.exports.docx_memo import Memo  # noqa: E402
from shared.exports.pdf import to_pdf  # noqa: E402
from shared.exports.pptx_deck import Deck  # noqa: E402
from shared.recalc import recalculate  # noqa: E402
from three_statement.report import FC, scenario_runs  # noqa: E402

LIMITS = [
    "Exit assumed as a single flow at the end of the hold — no dividend "
    "recaps or interim distributions (IRR therefore equals MOIC^(1/n)−1, "
    "which the Checks sheet verifies).",
    "Interest on beginning-of-period balances; no interest income on "
    "accumulated cash — both mildly conservative, both documented.",
    "Cash taxes take no benefit in loss years (no NOL carryforward).",
    "The revolver funds shortfalls against mandatory amortization at a "
    "spread — watch its balance in weak scenarios; persistent usage is the "
    "model telling you the structure doesn't work.",
    "Management incentives, monitoring fees and minimum cash operating "
    "requirements are not modeled.",
]


def _probes():
    p = ["Checks!B9", CHECKS_PASS_CELL]
    p += [f"Deal!B{r}" for r in DEAL.values()]
    p += [f"Returns!B{r}" for r in R.values() if r < R["flows"] or
          R["br_entry"] <= r <= R["br_exit"]]
    for c in FC:
        p += [f"LBO!{c}{L[k]}" for k in
              ("fcf", "rv_end", "sr_end", "sub_end", "cash_end", "debt_end",
               "net_debt", "ebitda")]
    for i in range(5):
        r = R["sens_first"] + i
        p += [f"Returns!B{r}"] + [f"Returns!{chr(ord('C') + k)}{r}"
                                  for k in range(3)]
    return p


def extract(xlsx) -> dict:
    res = recalculate(xlsx, probe_cells=_probes())
    if not res.ok:
        raise RuntimeError(f"{xlsx}: formula errors — refusing to report.\n"
                           + res.summary())
    v = res.values
    if v[CHECKS_PASS_CELL] != "PASS" or v["Checks!B9"] != "PASS":
        raise RuntimeError(f"{xlsx}: integrity checks FAIL — refusing to "
                           "report on a broken model.")
    wb = load_workbook(xlsx)
    title = wb["Assumptions"]["A1"].value or ""
    m = re.match(r"^(?P<name>.+?) \((?P<ticker>[A-Z.\-]+)\)", title)
    name, ticker = (m["name"], m["ticker"]) if m else (title, "?")
    deal = {k: v[f"Deal!B{r}"] for k, r in DEAL.items()}
    ret = {k: v[f"Returns!B{r}"] for k, r in R.items()
           if f"Returns!B{r}" in v}
    lbo = {k: [v[f"LBO!{c}{L[k]}"] for c in FC] for k in
           ("fcf", "rv_end", "sr_end", "sub_end", "cash_end", "debt_end",
            "net_debt", "ebitda")}
    sens = dict(
        mults=[v[f"Returns!B{R['sens_first'] + i}"] for i in range(5)],
        grid=[[v[f"Returns!{chr(ord('C') + k)}{R['sens_first'] + i}"]
               for k in range(3)] for i in range(5)],
    )
    return dict(name=name, ticker=ticker, deal=deal, ret=ret, lbo=lbo,
                sens=sens)


def leverage_rerun(xlsx, leverage_points=((2.5, 1.0), (3.5, 1.5), (4.5, 1.5)),
                   exit_deltas=(-1.0, 0.0, 1.0)) -> list:
    """Leverage x exit-multiple IRR table by re-running the real model."""
    rows = []
    probes = [f"Returns!B{R['irr']}", CHECKS_PASS_CELL,
              f"Deal!B{DEAL['exit_mult']}"]
    with tempfile.TemporaryDirectory() as tmp:
        for sr_x, sub_x in leverage_points:
            irr_row = []
            for d_exit in exit_deltas:
                p = Path(tmp) / f"l{sr_x}_{d_exit}.xlsx"
                shutil.copy(xlsx, p)
                wb = load_workbook(p)
                deal = wb["Deal"]
                deal[f"B{DEAL['senior_x']}"] = sr_x
                deal[f"B{DEAL['sub_x']}"] = sub_x
                # exit delta applies off the workbook's current exit input
                cur = wb["Deal"][f"B{DEAL['exit_mult']}"].value
                deal[f"B{DEAL['exit_mult']}"] = float(cur) + d_exit
                wb.save(p)
                res = recalculate(p, probe_cells=probes)
                ok = res.values[CHECKS_PASS_CELL] == "PASS"
                irr_row.append(res.values[f"Returns!B{R['irr']}"]
                               if ok else None)
            rows.append(((sr_x, sub_x), irr_row))
    return rows


def _sens_table(sens):
    rows = [["Exit mult \\ exit yr", "Year 3", "Year 4", "Year 5"]]
    for m, grid in zip(sens["mults"], sens["grid"]):
        rows.append([f"{m:.1f}x"] + [f"{x:.1%}" for x in grid])
    return rows


def _lev_table(lev_rows, base_exit):
    hdr = ["Leverage \\ exit"] + [f"{base_exit + d:.1f}x"
                                  for d in (-1.0, 0.0, 1.0)]
    rows = [hdr]
    for (sr, sub), irrs in lev_rows:
        rows.append([f"{sr + sub:.1f}x ({sr:.1f}sr+{sub:.1f}sub)"]
                    + [f"{x:.1%}" if x is not None else "breaks" for x in irrs])
    return rows


def _su_table(deal):
    return [["Sources & uses", "$mm"],
            ["Senior debt", f"{deal['senior_amt']:,.0f}"],
            ["Subordinated debt", f"{deal['sub_amt']:,.0f}"],
            ["Sponsor equity", f"{deal['s_equity']:,.0f}"],
            ["Total sources", f"{deal['s_tot']:,.0f}"],
            ["Enterprise value", f"{deal['entry_ev']:,.0f}"],
            ["Fees", f"{deal['fees']:,.0f}"],
            ["Total uses", f"{deal['u_tot']:,.0f}"]]


def build_reports(xlsx, outdir=None):
    xlsx = Path(xlsx)
    outdir = Path(outdir) if outdir else xlsx.parent
    d = extract(xlsx)
    deal, ret, lbo = d["deal"], d["ret"], d["lbo"]
    scen = scenario_runs(xlsx, extra_probes=[f"Returns!B{R['irr']}",
                                             f"Returns!B{R['moic']}",
                                             CHECKS_PASS_CELL])
    lev = leverage_rerun(xlsx)

    tmp = Path(tempfile.mkdtemp())
    years = [f"Year {t}" for t in range(1, 6)]
    png_debt = charts.debt_paydown(
        years, {"Senior": lbo["sr_end"], "Subordinated": lbo["sub_end"],
                "Revolver": lbo["rv_end"]}, tmp / "debt.png")
    png_bridge = charts.waterfall(
        [("Entry equity", ret["br_entry"], "total"),
         ("EBITDA growth", ret["br_growth"], "delta"),
         ("Multiple", ret["br_mult"], "delta"),
         ("Deleveraging", ret["br_delev"], "delta"),
         ("Fees & other", ret["br_fees"], "delta"),
         ("Exit equity", ret["br_exit"], "total")],
        tmp / "bridge.png")

    lev_total = deal["senior_x"] + deal["sub_x"]
    max_rv = max(lbo["rv_end"])
    summary = [
        f"Entry at {deal['entry_mult']:.1f}x LTM EBITDA "
        f"(${deal['entry_ev']:,.0f}mm EV), {lev_total:.1f}x total leverage, "
        f"${deal['s_equity']:,.0f}mm sponsor equity.",
        f"5-year IRR {ret['irr']:.1%}, MOIC {ret['moic']:.2f}x at a flat "
        f"{deal['exit_mult']:.1f}x exit.",
        f"Debt reduces ${lbo['debt_end'][0]:,.0f}mm → "
        f"${lbo['debt_end'][-1]:,.0f}mm through mandatory amort and full "
        "cash sweep.",
        ("Revolver undrawn — FCF covers debt service throughout."
         if max_rv <= 0.01 else
         f"Revolver peaks at ${max_rv:,.0f}mm — FCF does not fully cover "
         "mandatory amort; watch this in weak scenarios."),
        "Operating forecast is the linked three-statement model — the "
        "scenario toggle reprices the IRR.",
    ]
    scen_rows = [["Scenario", "IRR", "MOIC", "Checks"]]
    for lbl in ("Bear", "Base", "Bull"):
        s = scen[lbl]
        scen_rows.append([lbl, f"{s[f'Returns!B{R['irr']}']:.1%}",
                          f"{s[f'Returns!B{R['moic']}']:.2f}x",
                          s[CHECKS_PASS_CELL]])

    footer = (f"{d['ticker']} — LBO analysis | modeling-suite | $mm | "
              "AI-assembled from the recalculated model — verify before use")
    deck = Deck(footer)
    deck.title_slide(f"{d['name']} — LBO Analysis",
                     f"{deal['entry_mult']:.1f}x entry, {lev_total:.1f}x "
                     f"leverage, 5-year hold | {d['ticker']}",
                     "Live Excel model: blue = input, black = formula, "
                     "green = cross-sheet link")
    s = deck.content_slide("Executive summary — returns")
    deck.add_bullets(s, summary, 0.5, 1.3, 6.6, 5.4, size=12)
    deck.add_table(s, _su_table(deal), 7.6, 1.5, 5.2, first_col_w=2.6)

    s = deck.content_slide("Deleveraging")
    deck.add_image(s, png_debt, 0.5, 1.4, 7.4)
    deck.add_bullets(s, [
        f"FCF for debt service grows ${lbo['fcf'][0]:,.0f}mm → "
        f"${lbo['fcf'][-1]:,.0f}mm.",
        "Sweep order: revolver, senior, subordinated.",
        f"Net debt at exit ${lbo['net_debt'][-1]:,.0f}mm.",
    ], 8.2, 1.6, 4.6, 4.0, size=11)

    s = deck.content_slide("Value-creation bridge")
    deck.add_image(s, png_bridge, 0.7, 1.3, 8.2)
    deck.add_bullets(s, [
        "Flat exit multiple by design — returns must come from EBITDA "
        "growth and deleveraging, not multiple expansion.",
    ], 9.2, 1.8, 3.6, 3.0, size=10)

    s = deck.content_slide("Sensitivities & scenarios")
    deck.add_table(s, _sens_table(d["sens"]), 0.5, 1.5, 5.2, first_col_w=1.8)
    deck.add_table(s, _lev_table(lev, deal["exit_mult"]), 6.2, 1.5, 6.4,
                   first_col_w=2.4)
    deck.add_table(s, scen_rows, 6.2, 3.7, 6.4)
    deck.add_bullets(s, [
        "Left: live in-workbook grid (exit multiple × exit year). Right: "
        "leverage × exit multiple, computed by re-running the full model "
        "per leverage point — a live grid cannot honestly re-solve the "
        "sweep path.",
    ], 0.5, 4.6, 5.4, 2.2, size=10)

    s = deck.content_slide("Risks & limitations")
    deck.add_bullets(s, LIMITS, 0.5, 1.4, 12.0, 5.2, size=12)
    pptx_path = deck.save(outdir / f"{xlsx.stem}_deck.pptx")

    memo = Memo(f"{d['name']} ({d['ticker']}) — LBO Analysis Memo",
                "Leveraged buyout returns on the linked three-statement "
                "forecast | modeling-suite")
    memo.heading("Executive summary")
    memo.bullets(summary)
    memo.table(_su_table(deal))
    memo.heading("Deleveraging")
    memo.image(png_debt)
    memo.heading("Value creation")
    memo.image(png_bridge)
    memo.heading("Sensitivities & scenarios")
    memo.table(_sens_table(d["sens"]))
    memo.table(_lev_table(lev, deal["exit_mult"]))
    memo.table(scen_rows)
    memo.heading("Risks & limitations")
    memo.bullets(LIMITS)
    memo.para("Generated by modeling-suite from the recalculated workbook — "
              "figures cannot diverge from the model.", small=True)
    docx_path = memo.save(outdir / f"{xlsx.stem}_memo.docx")

    pdf_path = to_pdf(pptx_path, outdir)
    shutil.rmtree(tmp, ignore_errors=True)
    return dict(pptx=pptx_path, docx=docx_path, pdf=pdf_path)


def main():
    import argparse
    ap = argparse.ArgumentParser(description="Build deck/memo/PDF from an "
                                 "LBO workbook.")
    ap.add_argument("workbook")
    ap.add_argument("--outdir")
    args = ap.parse_args()
    for k, p in build_reports(args.workbook, args.outdir).items():
        print(f"{k}: {p}")


if __name__ == "__main__":
    main()
