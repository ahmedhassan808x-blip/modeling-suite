"""
Merger (M&A) model builder — accretion / dilution.

Five sheets:
  Acquirer / Target   compact earnings forecasts (3yr actuals blue, 5yr
                      driver-based forecast) — deliberately simpler than the
                      full three-statement build: a merger consequences model
                      needs clean EPS paths, not revolver mechanics
  Deal                offer terms, consideration mix, exchange ratio, fees,
                      financing, and purchase price allocation basics
                      (intangibles amortized, goodwill not)
  ProForma            combined NI with after-tax synergies, incremental
                      financing cost and intangible amortization; pro forma
                      EPS vs. standalone; breakeven synergies per year;
                      synergies x premium sensitivity (live)
  ChecksMA            mix sums to 100%, sources = uses, goodwill >= 0,
                      Y1 accretion re-derived independently — PASS cell

Conventions as suite-wide: blue/black/green enforced, costs negative,
no circular references, recalc-gate verified.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402

from shared.excel_utils import (  # noqa: E402
    BLUE, BOLD, GRAY, HDR_FILL, HDR_FONT, INPUT_HL, THIN_TOP,
    FMT_PCT, FMT_PS, FMT_USD, FMT_X, col_letter, set_cell, set_widths,
    sheet_title, year_headers,
)
from three_statement.assumptions import Assumptions  # noqa: E402

N_HIST, N_FC = 3, 5
WIDTHS = {"A": 42, "B": 14, **{col_letter(i): 12 for i in range(8)}}

CO = dict(growth=4, margin=5, da_pct=6, tax=7,
          rev=10, ebitda=11, ebitda_m=12, da=13, ebit=14, other=15,
          ebt=16, taxes=17, ni=18, shares=20, eps=21)
DEAL = dict(acq_price=5, acq_shares=6, tgt_price=8, tgt_shares=9,
            premium=10, offer_ps=11, offer_val=12, xratio=13,
            mix_stock=15, mix_cash=16, mix_debt=17, fees_pct=18, fees=19,
            stock_cons=21, new_shares=22, cash_cons=23, debt_cons=24,
            debt_rate=26, cash_yield=27, incr_int=28,
            tgt_book=31, prem_book=32, intang_pct=33, intang=34,
            amort_years=35, amort=36, goodwill=37,
            syn=40)
PF = dict(acq_ni=5, tgt_ni=6, phase=7, syn=8, int_=9, amort=10, adj_pre=11,
          adj_tax=12, adj_at=13, pf_ni=14, acq_sh=16, new_sh=17, pf_sh=18,
          acq_eps=20, pf_eps=21, acc_ps=22, acc_pct=23, breakeven=25,
          sens_hdr=28, sens_first=29)
CHECKS_PASS_CELL = "ChecksMA!B9"


def _fc_cols():
    return [col_letter(N_HIST + i) for i in range(N_FC)]


def _hist_cols():
    return [col_letter(i) for i in range(N_HIST)]


def build_company(ws, data, role):
    """Compact earnings forecast: drivers (blue) + IS to EPS."""
    hist, fc = _hist_cols(), _fc_cols()
    a = Assumptions.derive(data)
    sheet_title(ws, f"{role}: {data['name']} ({data['ticker']}) — Earnings "
                "forecast ($mm)",
                "Compact by design — a merger consequences model needs clean "
                "EPS paths. Net interest & other held flat at last actual.")
    year_headers(ws, 3, data["years"], N_FC)

    drivers = [("growth", "Revenue growth", a.rev_growth),
               ("margin", "EBITDA margin",
                [g - o for g, o in zip(a.gross_margin, a.opex_pct)]),
               ("da_pct", "D&A (% of revenue)", a.da_pct),
               ("tax", "Tax rate", a.tax_rate)]
    for key, label, vals in drivers:
        set_cell(ws, f"A{CO[key]}", label, GRAY)
        for j, c in enumerate(fc):
            set_cell(ws, f"{c}{CO[key]}", round(vals[j], 4), BLUE, FMT_PCT,
                     fill=INPUT_HL)

    labels = [("rev", "Revenue"), ("ebitda", "EBITDA"),
              ("ebitda_m", "  % margin"), ("da", "D&A"), ("ebit", "EBIT"),
              ("other", "Net interest & other (held flat)"),
              ("ebt", "Pre-tax income"), ("taxes", "Income taxes"),
              ("ni", "Net income"), ("shares", "Shares outstanding (mm)"),
              ("eps", "EPS ($)")]
    for key, label in labels:
        set_cell(ws, f"A{CO[key]}", label, None,
                 bold=key in ("ebitda", "ebit", "ni", "eps"))

    isl = data["is"]
    for i, c in enumerate(hist):
        ebitda = isl["ebit"][i] + isl["dna"][i]
        set_cell(ws, f"{c}{CO['rev']}", isl["revenue"][i], BLUE, FMT_USD)
        set_cell(ws, f"{c}{CO['ebitda']}", ebitda, BLUE, FMT_USD)
        set_cell(ws, f"{c}{CO['da']}", -isl["dna"][i], BLUE, FMT_USD)
        set_cell(ws, f"{c}{CO['other']}", isl["pretax"][i] - isl["ebit"][i],
                 BLUE, FMT_USD)
        set_cell(ws, f"{c}{CO['taxes']}", -isl["tax"][i], BLUE, FMT_USD)
    set_cell(ws, f"B{CO['shares']}", data["shares_mm"], BLUE, FMT_USD)

    for j, c in enumerate(fc):
        p = col_letter(N_HIST + j - 1)
        set_cell(ws, f"{c}{CO['rev']}",
                 f"={p}{CO['rev']}*(1+{c}{CO['growth']})", None, FMT_USD)
        set_cell(ws, f"{c}{CO['ebitda']}",
                 f"={c}{CO['rev']}*{c}{CO['margin']}", None, FMT_USD)
        set_cell(ws, f"{c}{CO['da']}",
                 f"=-{c}{CO['rev']}*{c}{CO['da_pct']}", None, FMT_USD)
        set_cell(ws, f"{c}{CO['other']}", f"={p}{CO['other']}", None, FMT_USD)
        set_cell(ws, f"{c}{CO['taxes']}",
                 f"=-{c}{CO['ebt']}*{c}{CO['tax']}", None, FMT_USD)
    for c in hist + fc:
        set_cell(ws, f"{c}{CO['ebitda_m']}",
                 f"={c}{CO['ebitda']}/{c}{CO['rev']}", GRAY, FMT_PCT)
        set_cell(ws, f"{c}{CO['ebit']}",
                 f"={c}{CO['ebitda']}+{c}{CO['da']}", None, FMT_USD,
                 bold=True)
        set_cell(ws, f"{c}{CO['ebt']}",
                 f"={c}{CO['ebit']}+{c}{CO['other']}", None, FMT_USD)
        set_cell(ws, f"{c}{CO['ni']}",
                 f"={c}{CO['ebt']}+{c}{CO['taxes']}", None, FMT_USD,
                 bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{CO['eps']}",
                 f"={c}{CO['ni']}/$B${CO['shares']}", None, FMT_PS, bold=True)
    set_widths(ws, WIDTHS)


def build_deal(ws, acq, tgt, m):
    tgt_book = sum(tgt["bs"][k][-1] for k in
                   ("cs_apic", "re", "other_eq", "minority"))
    syn_seed = m.seed_synergies(tgt)
    sheet_title(ws, f"Deal: {acq['ticker']} acquires {tgt['ticker']} ($mm)",
                "Offer terms, financing and purchase price allocation "
                "basics. Cost synergies only in the base case — revenue "
                "synergies excluded by discipline.")
    rows = [
        ("acq_price", "Acquirer share price ($)", acq["price"], FMT_PS, True),
        ("acq_shares", "Acquirer shares (mm)", acq["shares_mm"], FMT_USD,
         True),
        ("tgt_price", "Target share price ($)", tgt["price"], FMT_PS, True),
        ("tgt_shares", "Target shares (mm)", tgt["shares_mm"], FMT_USD, True),
        ("premium", "Acquisition premium", m.premium, FMT_PCT, True),
        ("offer_ps", "Offer per share ($)",
         f"=B{DEAL['tgt_price']}*(1+B{DEAL['premium']})", FMT_PS, False),
        ("offer_val", "Offer value (equity)",
         f"=B{DEAL['offer_ps']}*B{DEAL['tgt_shares']}", FMT_USD, False),
        ("xratio", "Exchange ratio (stock portion, tgt sh -> acq sh)",
         f"=B{DEAL['offer_ps']}/B{DEAL['acq_price']}", "0.000x", False),
        ("mix_stock", "Consideration — stock", m.mix_stock, FMT_PCT, True),
        ("mix_cash", "Consideration — balance-sheet cash", m.mix_cash,
         FMT_PCT, True),
        ("mix_debt", "Consideration — new debt", m.mix_debt, FMT_PCT, True),
        ("fees_pct", "Transaction fees (% of offer, debt-financed)",
         m.fees_pct, FMT_PCT, True),
        ("fees", "Transaction fees",
         f"=B{DEAL['fees_pct']}*B{DEAL['offer_val']}", FMT_USD, False),
        ("stock_cons", "Stock consideration",
         f"=B{DEAL['mix_stock']}*B{DEAL['offer_val']}", FMT_USD, False),
        ("new_shares", "New acquirer shares issued (mm)",
         f"=B{DEAL['stock_cons']}/B{DEAL['acq_price']}", FMT_USD, False),
        ("cash_cons", "Cash consideration",
         f"=B{DEAL['mix_cash']}*B{DEAL['offer_val']}", FMT_USD, False),
        ("debt_cons", "New debt (incl. fees)",
         f"=B{DEAL['mix_debt']}*B{DEAL['offer_val']}+B{DEAL['fees']}",
         FMT_USD, False),
        ("debt_rate", "Rate on new debt", m.debt_rate, FMT_PCT, True),
        ("cash_yield", "Foregone yield on cash used", m.cash_yield, FMT_PCT,
         True),
        ("incr_int", "Incremental pre-tax financing cost / yr",
         f"=B{DEAL['debt_cons']}*B{DEAL['debt_rate']}"
         f"+B{DEAL['cash_cons']}*B{DEAL['cash_yield']}", FMT_USD, False),
        ("tgt_book", "Target book equity", round(tgt_book, 1), FMT_USD, True),
        ("prem_book", "Premium over book",
         f"=B{DEAL['offer_val']}-B{DEAL['tgt_book']}", FMT_USD, False),
        ("intang_pct", "Allocated to identifiable intangibles",
         m.intang_pct, FMT_PCT, True),
        ("intang", "Identifiable intangibles",
         f"=B{DEAL['prem_book']}*B{DEAL['intang_pct']}", FMT_USD, False),
        ("amort_years", "Intangible amortization (years, straight-line)",
         m.amort_years, "0", True),
        ("amort", "Incremental amortization / yr",
         f"=B{DEAL['intang']}/B{DEAL['amort_years']}", FMT_USD, False),
        ("goodwill", "Goodwill (not amortized)",
         f"=B{DEAL['prem_book']}-B{DEAL['intang']}", FMT_USD, False),
        ("syn", "Pre-tax run-rate COST synergies "
         "(seed: 10% of target EBITDA — placeholder, overwrite with a view)",
         syn_seed, FMT_USD, True),
    ]
    set_cell(ws, "A4", "Offer & consideration", None, bold=True)
    set_cell(ws, "A30", "Purchase price allocation (basics)", None, bold=True)
    set_cell(ws, "A39", "Synergies", None, bold=True)
    for key, label, val, fmt, is_input in rows:
        set_cell(ws, f"A{DEAL[key]}", label)
        c = set_cell(ws, f"B{DEAL[key]}", val, BLUE if is_input else None,
                     fmt)
        if is_input:
            c.fill = INPUT_HL
    set_widths(ws, WIDTHS)


def build_proforma(ws, acq, tgt, m):
    fc = _fc_cols()
    sheet_title(ws, f"Pro Forma: {acq['ticker']} + {tgt['ticker']} ($mm)",
                "Accretion/(dilution) vs. acquirer standalone EPS. New debt "
                "carried flat (no paydown modeled — documented limitation).")
    for t, c in enumerate(fc, start=1):
        set_cell(ws, f"{c}3", f"Year {t}", BOLD)

    labels = [
        ("acq_ni", "Acquirer net income", False),
        ("tgt_ni", "Target net income", False),
        ("phase", "Synergy phase-in (%)", False),
        ("syn", "(+) Pre-tax synergies", False),
        ("int_", "(–) Incremental financing cost", False),
        ("amort", "(–) Incremental intangible amortization", False),
        ("adj_pre", "Pre-tax adjustments", False),
        ("adj_tax", "Tax effect (acquirer rate)", False),
        ("adj_at", "After-tax adjustments", False),
        ("pf_ni", "Pro forma net income", True),
        ("acq_sh", "Acquirer shares (mm)", False),
        ("new_sh", "New shares issued (mm)", False),
        ("pf_sh", "Pro forma shares (mm)", True),
        ("acq_eps", "Acquirer standalone EPS ($)", False),
        ("pf_eps", "Pro forma EPS ($)", True),
        ("acc_ps", "Accretion / (dilution) per share ($)", False),
        ("acc_pct", "Accretion / (dilution) %", True),
        ("breakeven", "Breakeven pre-tax run-rate synergies", False),
    ]
    for key, label, bold in labels:
        set_cell(ws, f"A{PF[key]}", label, None, bold=bold)

    for j, c in enumerate(fc):
        set_cell(ws, f"{c}{PF['acq_ni']}", f"=Acquirer!{c}{CO['ni']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{PF['tgt_ni']}", f"=Target!{c}{CO['ni']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{PF['phase']}", m.phase_in[j], BLUE, FMT_PCT,
                 fill=INPUT_HL)
        set_cell(ws, f"{c}{PF['syn']}",
                 f"=Deal!$B${DEAL['syn']}*{c}{PF['phase']}", None, FMT_USD)
        set_cell(ws, f"{c}{PF['int_']}", f"=-Deal!$B${DEAL['incr_int']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{PF['amort']}", f"=-Deal!$B${DEAL['amort']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{PF['adj_pre']}",
                 f"=SUM({c}{PF['syn']}:{c}{PF['amort']})", None, FMT_USD)
        set_cell(ws, f"{c}{PF['adj_tax']}",
                 f"=-{c}{PF['adj_pre']}*Acquirer!{c}{CO['tax']}", None,
                 FMT_USD)
        set_cell(ws, f"{c}{PF['adj_at']}",
                 f"={c}{PF['adj_pre']}+{c}{PF['adj_tax']}", None, FMT_USD)
        set_cell(ws, f"{c}{PF['pf_ni']}",
                 f"={c}{PF['acq_ni']}+{c}{PF['tgt_ni']}+{c}{PF['adj_at']}",
                 None, FMT_USD, bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{PF['acq_sh']}", f"=Deal!$B${DEAL['acq_shares']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{PF['new_sh']}", f"=Deal!$B${DEAL['new_shares']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{PF['pf_sh']}",
                 f"={c}{PF['acq_sh']}+{c}{PF['new_sh']}", None, FMT_USD,
                 bold=True)
        set_cell(ws, f"{c}{PF['acq_eps']}", f"=Acquirer!{c}{CO['eps']}",
                 None, FMT_PS)
        set_cell(ws, f"{c}{PF['pf_eps']}",
                 f"={c}{PF['pf_ni']}/{c}{PF['pf_sh']}", None, FMT_PS,
                 bold=True)
        set_cell(ws, f"{c}{PF['acc_ps']}",
                 f"={c}{PF['pf_eps']}-{c}{PF['acq_eps']}", None, FMT_PS)
        set_cell(ws, f"{c}{PF['acc_pct']}",
                 f"={c}{PF['pf_eps']}/{c}{PF['acq_eps']}-1", None, FMT_PCT,
                 bold=True, border=THIN_TOP)
        # synergies s.t. PF EPS == standalone EPS:
        # s*phase = (eps*pf_sh - acq_ni - tgt_ni)/(1-tax) + int + amort
        set_cell(ws, f"{c}{PF['breakeven']}",
                 f"=(({c}{PF['acq_eps']}*{c}{PF['pf_sh']}-{c}{PF['acq_ni']}"
                 f"-{c}{PF['tgt_ni']})/(1-Acquirer!{c}{CO['tax']})"
                 f"+Deal!$B${DEAL['incr_int']}+Deal!$B${DEAL['amort']})"
                 f"/{c}{PF['phase']}", None, FMT_USD)

    # Sensitivity: run-rate synergies (rows) x premium (cols) -> Y1 accretion
    set_cell(ws, f"A{PF['sens_hdr'] - 1}",
             "Sensitivity — Year 1 accretion/(dilution): synergies × premium",
             None, bold=True)
    set_cell(ws, f"B{PF['sens_hdr']}", "Syn \\ premium", BOLD)
    f1 = fc[0]
    d = "Deal!$B$"
    for k in range(5):
        set_cell(ws, f"{chr(ord('C') + k)}{PF['sens_hdr']}",
                 f"={d}{DEAL['premium']}+{k - 2}*0.05", None, FMT_PCT)
    for i in range(5):
        r = PF["sens_first"] + i
        set_cell(ws, f"B{r}", f"={d}{DEAL['syn']}*(1+{i - 2}*0.5)", None,
                 FMT_USD)
        for k in range(5):
            sc = chr(ord("C") + k)
            prem, syn = f"{sc}${PF['sens_hdr']}", f"$B{r}"
            offer = f"({d}{DEAL['tgt_price']}*(1+{prem})*{d}{DEAL['tgt_shares']})"
            newsh = f"({d}{DEAL['mix_stock']}*{offer}/{d}{DEAL['acq_price']})"
            debt = f"({d}{DEAL['mix_debt']}*{offer}+{d}{DEAL['fees_pct']}*{offer})"
            cash = f"({d}{DEAL['mix_cash']}*{offer})"
            intr = f"({debt}*{d}{DEAL['debt_rate']}+{cash}*{d}{DEAL['cash_yield']})"
            amort = (f"(({offer}-{d}{DEAL['tgt_book']})*{d}{DEAL['intang_pct']}"
                     f"/{d}{DEAL['amort_years']})")
            adj = (f"(({syn}*{f1}${PF['phase']}-{intr}-{amort})"
                   f"*(1-Acquirer!{f1}${CO['tax']}))")
            pf_ni = f"({f1}${PF['acq_ni']}+{f1}${PF['tgt_ni']}+{adj})"
            pf_eps = f"({pf_ni}/({f1}${PF['acq_sh']}+{newsh}))"
            set_cell(ws, f"{sc}{r}", f"={pf_eps}/{f1}${PF['acq_eps']}-1",
                     None, FMT_PCT)
    set_widths(ws, WIDTHS)


def build_checks_ma(ws, acq, tgt):
    f1 = _fc_cols()[0]
    sheet_title(ws, f"{acq['ticker']} / {tgt['ticker']} — M&A Integrity "
                "Checks", "Everything zero and B9 = PASS; asserted by the "
                "recalc gate.")
    set_cell(ws, "A4", "Consideration mix sums to 100%")
    set_cell(ws, "B4", f"=Deal!B{DEAL['mix_stock']}+Deal!B{DEAL['mix_cash']}"
             f"+Deal!B{DEAL['mix_debt']}-1", None, "0.000")
    set_cell(ws, "A5", "Sources − uses (stock+cash+debt − offer − fees)")
    set_cell(ws, "B5", f"=Deal!B{DEAL['stock_cons']}+Deal!B{DEAL['cash_cons']}"
             f"+Deal!B{DEAL['debt_cons']}-Deal!B{DEAL['offer_val']}"
             f"-Deal!B{DEAL['fees']}", None, "0.000")
    set_cell(ws, "A6", "Goodwill non-negative (0 = ok)")
    set_cell(ws, "B6", f"=IF(Deal!B{DEAL['goodwill']}>=-0.001,0,1)", None, "0")
    set_cell(ws, "A7", "Y1 accretion: sensitivity center == ProForma "
             "(independent re-derivation)")
    set_cell(ws, "B7", f"=ProForma!E{PF['sens_first'] + 2}"
             f"-ProForma!{f1}{PF['acc_pct']}", None, "0.0000")
    set_cell(ws, "A9", "ALL M&A CHECKS", None, bold=True)
    cell = set_cell(ws, "B9",
                    '=IF(ABS(B4)+ABS(B5)+B6+ABS(B7)<0.005,"PASS","FAIL")')
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    set_widths(ws, WIDTHS)


def build_merger_model(acq_data, tgt_data, m, out_path):
    for d, role in ((acq_data, "Acquirer"), (tgt_data, "Target")):
        if not d.get("price") or not d.get("shares_mm"):
            raise ValueError(f"{role} {d['ticker']}: no market data — "
                             "cannot build a merger model.")
    wb = Workbook()
    ws = wb.active
    ws.title = "Acquirer"
    build_company(ws, acq_data, "Acquirer")
    build_company(wb.create_sheet("Target"), tgt_data, "Target")
    build_deal(wb.create_sheet("Deal"), acq_data, tgt_data, m)
    build_proforma(wb.create_sheet("ProForma"), acq_data, tgt_data, m)
    build_checks_ma(wb.create_sheet("ChecksMA"), acq_data, tgt_data)
    wb.save(out_path)
    return out_path
