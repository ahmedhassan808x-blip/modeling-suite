"""
Shared Excel-writing utility for all modeling-suite workbooks.

Conventions (industry standard, matching analyst-toolkit):
  blue font  = hardcoded input      black = formula      green = cross-sheet link
  negatives in parentheses, units declared in headers, sources on every input block.

All model calculations are written as live Excel formulas — never pre-computed
values — so the model recalculates when an analyst changes an input.

set_cell() enforces the color convention automatically: a formula containing a
sheet reference ("!") is colored green unless explicitly overridden, so the
convention cannot silently drift as models grow.
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Fonts / fills (Arial 10, matching analyst-toolkit)
BLUE = Font(name="Arial", size=10, color="0000FF")            # hardcoded input
BLACK = Font(name="Arial", size=10, color="000000")           # formula
GREEN = Font(name="Arial", size=10, color="008000")           # cross-sheet link
BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=12, bold=True)
GRAY = Font(name="Arial", size=9, color="808080", italic=True)  # notes / memo labels

# Navy & steel palette (Phase 3 exports share these)
NAVY = "1F3864"
STEEL = "8EAADB"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
INPUT_HL = PatternFill("solid", fgColor="FFFF99")   # yellow = challenge-first inputs
THIN_TOP = Border(top=Side(style="thin"))
DOUBLE_TOP = Border(top=Side(style="double"))

# Number formats — model works in $mm
FMT_USD = '#,##0;(#,##0);"-"'
FMT_USD1 = '#,##0.0;(#,##0.0);"-"'
FMT_PCT = "0.0%"
FMT_X = "0.0x"
FMT_PS = '$#,##0.00;($#,##0.00);"-"'
FMT_DAYS = '#,##0.0 "d"'
MM = 1e6


def set_cell(ws, cell, value, font=None, fmt=None, bold=False, fill=None,
             border=None):
    """Write one cell, enforcing the color convention.

    Explicit font always wins. Otherwise: cross-sheet formulas go green,
    same-sheet formulas black, and plain values black (pass font=BLUE for
    inputs — inputs must be deliberate, never a default).
    """
    c = ws[cell]
    c.value = value
    if font is None:
        if isinstance(value, str) and value.startswith("="):
            font = GREEN if "!" in value else BLACK
        else:
            font = BLACK
    c.font = BOLD if bold else font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if border:
        c.border = border
    return c


def set_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def col_letter(idx0: int, first_col: int = 3) -> str:
    """0-based data-column index -> letter. Data starts in column C by default."""
    return get_column_letter(first_col + idx0)


def year_headers(ws, row, years_hist, n_forecast, first_col=3):
    """FY2023A ... FY2027E header row, bold, aligned across all sheets."""
    for j, yr in enumerate(years_hist):
        set_cell(ws, f"{col_letter(j, first_col)}{row}", f"FY{yr}A", BOLD)
    last = years_hist[-1]
    for j in range(n_forecast):
        set_cell(ws, f"{col_letter(len(years_hist) + j, first_col)}{row}",
                 f"FY{last + 1 + j}E", BOLD)


def sheet_title(ws, title, subtitle):
    set_cell(ws, "A1", title, TITLE)
    set_cell(ws, "A2", subtitle, GRAY)
