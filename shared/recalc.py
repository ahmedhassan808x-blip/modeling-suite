"""
Automated recalculation gate for every workbook this suite produces.

How it works: openpyxl writes formulas WITHOUT cached results, so when
LibreOffice opens the file headlessly and re-saves it, every formula must be
genuinely computed by a real spreadsheet engine. We then reload the converted
file values-only and scan every cell for Excel error strings. A workbook passes
only with zero errors — this is the "delivered with zero formula errors"
guarantee, verified, not asserted.

Design constraint this imposes on all models: no circular references
(e.g. interest on average debt balances), because LibreOffice cannot cleanly
recalculate iterative models. Interest is computed on beginning-of-period
balances instead — a legitimate banking convention, documented per model.
"""

import shutil
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

EXCEL_ERRORS = {"#REF!", "#NAME?", "#DIV/0!", "#VALUE!", "#N/A", "#NUM!", "#NULL!"}

SOFFICE_CANDIDATES = [
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    shutil.which("soffice") or "",
]


class RecalcError(RuntimeError):
    pass


def find_soffice() -> str:
    for c in SOFFICE_CANDIDATES:
        if c and Path(c).exists():
            return c
    raise RecalcError(
        "LibreOffice not found — the recalc gate cannot run. "
        "Install it (https://www.libreoffice.org) or brew install --cask libreoffice."
    )


@dataclass
class RecalcResult:
    ok: bool
    errors: list = field(default_factory=list)   # (sheet, cell, error_value)
    values: dict = field(default_factory=dict)   # "Sheet!A1" -> recalculated value

    def summary(self) -> str:
        if self.ok:
            return "recalc PASS — zero formula errors"
        lines = [f"recalc FAIL — {len(self.errors)} formula error(s):"]
        lines += [f"  {s}!{c} = {v}" for s, c, v in self.errors[:20]]
        return "\n".join(lines)


def recalculate(xlsx_path, probe_cells=None) -> RecalcResult:
    """Recalculate a workbook via headless LibreOffice; scan for errors.

    probe_cells: optional list of "Sheet!A1" refs whose recalculated values
    are returned for assertions (e.g. balance checks must equal zero).
    """
    xlsx_path = Path(xlsx_path).resolve()
    if not xlsx_path.exists():
        raise RecalcError(f"No such workbook: {xlsx_path}")
    soffice = find_soffice()

    with tempfile.TemporaryDirectory() as tmp:
        # Isolated profile: avoids lock conflicts with a running LibreOffice GUI.
        profile = Path(tmp) / "profile"
        proc = subprocess.run(
            [soffice, "--headless", "--norestore",
             f"-env:UserInstallation=file://{profile}",
             "--convert-to", "xlsx", "--outdir", tmp, str(xlsx_path)],
            capture_output=True, text=True, timeout=180,
        )
        converted = Path(tmp) / xlsx_path.name
        if proc.returncode != 0 or not converted.exists():
            raise RecalcError(
                f"LibreOffice conversion failed (rc={proc.returncode}):\n"
                f"{proc.stdout}\n{proc.stderr}")

        wb = load_workbook(converted, data_only=True)  # cached = recalculated values
        errors, values = [], {}
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value in EXCEL_ERRORS:
                        errors.append((ws.title, c.coordinate, c.value))
        for ref in probe_cells or []:
            sheet, cell = ref.split("!")
            values[ref] = wb[sheet][cell].value

    return RecalcResult(ok=not errors, errors=errors, values=values)
