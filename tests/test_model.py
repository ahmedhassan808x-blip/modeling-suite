"""
The non-negotiable gate: every workbook must recalculate in a real
spreadsheet engine (headless LibreOffice) with zero formula errors, a
balanced balance sheet in every column, and Checks!B9 = PASS.

Also covered: revolver stress (forced draws must still balance), the
blue/black/green color convention, and hand-verified forecast arithmetic.
"""

import pytest
from openpyxl import Workbook, load_workbook

from shared.recalc import recalculate
from three_statement.assumptions import Assumptions
from three_statement.model_builder import BS, CF, DEBT, IS, build_model

HIST_FC_COLS = list("CDEFGHIJ")
FC_COLS = list("FGHIJ")


def _probes():
    p = [f"Checks!{c}5" for c in HIST_FC_COLS] + ["Checks!B9"]
    p += [f"CF!{c}{CF['end']}" for c in FC_COLS]
    p += [f"Debt!{c}{DEBT['rev_end']}" for c in FC_COLS]
    p += [f"IS!F{IS['rev']}", f"IS!F{IS['ni']}", f"BS!F{BS['re']}"]
    return p


def _build_and_recalc(data, assumptions, tmp_path, name):
    out = tmp_path / name
    build_model(data, assumptions, out)
    return recalculate(out, probe_cells=_probes())


def test_base_case_recalculates_clean(simpleco, tmp_path):
    a = Assumptions.derive(simpleco)
    res = _build_and_recalc(simpleco, a, tmp_path, "smpl.xlsx")
    assert res.ok, res.summary()
    assert res.values["Checks!B9"] == "PASS"
    for c in HIST_FC_COLS:  # balance sheet balances in EVERY column
        assert abs(res.values[f"Checks!{c}5"]) < 1e-3, f"BS out of balance in {c}"


def test_forecast_arithmetic_hand_checked(simpleco, tmp_path):
    """Y1 revenue = 1210 * (1 + trailing 10% growth) = 1331, exactly."""
    a = Assumptions.derive(simpleco)
    res = _build_and_recalc(simpleco, a, tmp_path, "smpl.xlsx")
    assert res.values[f"IS!F{IS['rev']}"] == pytest.approx(1331, rel=1e-6)
    # NI: rev 1331, GM 60% -> GP 798.6; opex 35% -> 465.85; D&A ~4.545% avg;
    # interest on beginning balances: 200*5.5% debt, 340*3% cash.
    ni = res.values[f"IS!F{IS['ni']}"]
    assert 180 < ni < 260, f"Y1 net income {ni} outside plausible band"
    # Retained earnings roll: prior RE + NI - dividends
    assert res.values[f"BS!F{BS['re']}"] > 680


def test_no_revolver_needed_in_base_case(simpleco, tmp_path):
    a = Assumptions.derive(simpleco)
    res = _build_and_recalc(simpleco, a, tmp_path, "smpl.xlsx")
    for c in FC_COLS:
        assert res.values[f"Debt!{c}{DEBT['rev_end']}"] == pytest.approx(0, abs=1e-6)


def test_revolver_stress_draws_and_still_balances(simpleco, tmp_path):
    """Starve the company of cash: heavy capex + full payout + high min cash.
    The revolver must draw (that's the plug working) and A = L + E must hold."""
    a = Assumptions.derive(simpleco)
    a.capex_pct = [0.60] * 5
    a.payout = [1.0] * 5
    a.min_cash_floor = 300.0
    res = _build_and_recalc(simpleco, a, tmp_path, "stress.xlsx")
    assert res.ok, res.summary()
    assert res.values["Checks!B9"] == "PASS"
    draws = [res.values[f"Debt!{c}{DEBT['rev_end']}"] for c in FC_COLS]
    assert max(draws) > 0, "stress case never drew the revolver"
    ends = [res.values[f"CF!{c}{CF['end']}"] for c in FC_COLS]
    assert all(e >= 300 - 0.01 for e in ends), "cash fell below the minimum target"


def test_color_convention_enforced(simpleco, tmp_path):
    a = Assumptions.derive(simpleco)
    out = tmp_path / "smpl.xlsx"
    build_model(simpleco, a, out)
    wb = load_workbook(out)
    assert wb["IS"][f"C{IS['rev']}"].font.color.rgb.endswith("0000FF")  # input
    assert wb["IS"][f"F{IS['rev']}"].font.color.rgb.endswith("008000")  # x-sheet
    assert wb["BS"][f"F{BS['gwi']}"].font.color.rgb.endswith("000000")  # formula
    for c in FC_COLS:  # forecast statements contain no pasted values
        for row in (IS["rev"], IS["ni"]):
            v = wb["IS"][f"{c}{row}"].value
            assert isinstance(v, str) and v.startswith("=")


def test_recalc_gate_catches_errors(tmp_path):
    wb = Workbook()
    wb.active["A1"] = "=1/0"
    wb.active["A2"] = "=NoSheet!B2"
    bad = tmp_path / "bad.xlsx"
    wb.save(bad)
    res = recalculate(bad)
    assert not res.ok and len(res.errors) >= 2
