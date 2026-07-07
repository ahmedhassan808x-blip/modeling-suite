"""
Scenario toggle: flipping one cell must repropagate through the entire
model — statements, revolver, checks — with bear < base < bull ordering
and a balanced balance sheet in every case.
"""

import pytest
from openpyxl import load_workbook

from shared.recalc import recalculate
from three_statement.assumptions import Assumptions
from three_statement.model_builder import IS, build_model


def _recalc_with_selector(simpleco, tmp_path, selector):
    a = Assumptions.derive(simpleco)
    out = tmp_path / f"scen{selector}.xlsx"
    build_model(simpleco, a, out)
    wb = load_workbook(out)            # flip the toggle like an analyst would
    wb["Assumptions"]["B3"] = selector
    wb.save(out)
    return recalculate(out, probe_cells=["Checks!B9", f"IS!J{IS['rev']}",
                                         f"IS!J{IS['ni']}"])


def test_scenarios_reprice_and_balance(simpleco, tmp_path):
    results = {s: _recalc_with_selector(simpleco, tmp_path, s)
               for s in (1, 2, 3)}
    for s, res in results.items():
        assert res.ok, f"scenario {s}: {res.summary()}"
        assert res.values["Checks!B9"] == "PASS", f"scenario {s} out of balance"
    rev = {s: r.values[f"IS!J{IS['rev']}"] for s, r in results.items()}
    ni = {s: r.values[f"IS!J{IS['ni']}"] for s, r in results.items()}
    assert rev[1] < rev[2] < rev[3], f"Y5 revenue ordering wrong: {rev}"
    assert ni[1] < ni[2] < ni[3], f"Y5 net income ordering wrong: {ni}"


def test_base_case_matches_derived_assumptions(simpleco, tmp_path):
    """Selector=2 (Base) must reproduce the pre-scenario Phase 1 numbers."""
    res = _recalc_with_selector(simpleco, tmp_path, 2)
    # Y5 revenue: 1210 compounding at growth fading 10% -> 3%
    a = Assumptions.derive(simpleco)
    expected = 1210.0
    for g in a.rev_growth:
        expected *= 1 + g
    assert res.values[f"IS!J{IS['rev']}"] == pytest.approx(expected, rel=1e-6)


def test_scenario_inputs_live_on_scenarios_sheet(simpleco, tmp_path):
    a = Assumptions.derive(simpleco)
    out = tmp_path / "s.xlsx"
    build_model(simpleco, a, out)
    wb = load_workbook(out)
    # Assumptions growth row is now a CHOOSE formula, not a hardcode
    v = wb["Assumptions"]["F6"].value
    assert isinstance(v, str) and v.startswith("=CHOOSE(")
    # ...and the Scenarios sheet holds blue inputs (bear growth < bull growth)
    assert wb["Scenarios"]["F6"].value < wb["Scenarios"]["F18"].value
    assert wb["Scenarios"]["F6"].font.color.rgb.endswith("0000FF")
