"""
Three-statement model builder — the foundation workbook of the suite.

Six sheets: Assumptions, IS, BS, CF, Debt, Checks. Historical columns are blue
hardcodes tied exactly to filings (via plug lines computed in the data layer);
forecast columns are 100% live Excel formulas. Nothing is pre-computed.

The full linkage, as taught and as checked:
  net income -> retained earnings; D&A/capex roll PP&E; working capital moves
  from days ratios; the revolver draws when cash would fall below the minimum
  and sweeps back down when there's excess; CF ending cash IS the balance
  sheet cash. The Checks sheet proves A = L + E in every column.

No circular references by design: interest accrues on BEGINNING-of-period
balances, so the recalc gate (LibreOffice, which cannot iterate) can verify
the workbook. See README for why this convention was chosen.
"""

from datetime import date

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from shared.excel_utils import (  # noqa: E402
    BLUE, BOLD, GRAY, GREEN, HDR_FILL, HDR_FONT, INPUT_HL, THIN_TOP,
    FMT_DAYS, FMT_PCT, FMT_USD, col_letter, set_cell, set_widths,
    sheet_title, year_headers,
)
from three_statement.scenarios import (  # noqa: E402
    BEAR, BULL, SCENARIO_DRIVERS, seed_scenarios,
)

N_FC = 5

# Row maps — single source of truth for every cross-sheet reference.
ASM = dict(growth=6, gm=7, opex=8, da=9, capex=10,
           dso=12, dio=13, dpo=14, oca=15, ocl=16, tax=18, payout=19)
SELECTOR = "$B$3"  # Assumptions scenario toggle: 1=Bear, 2=Base, 3=Bull
SCEN = {"Bear": dict(growth=6, gm=7, opex=8, capex=9),
        "Base": dict(growth=12, gm=13, opex=14, capex=15),
        "Bull": dict(growth=18, gm=19, opex=20, capex=21)}
IS = dict(rev=5, growth=6, cogs=7, gp=8, gm=9, opex=10, ebitda=11, ebitda_m=12,
          da=13, ebit=14, int_exp=15, int_inc=16, other=17, ebt=18, tax=19,
          etr=20, ni=21)
BS = dict(cash=5, ar=6, inv=7, oca=8, tca=9, ppe=10, gwi=11, onca=12, ta=13,
          ap=15, ocl=16, rev=17, ltd=18, oncl=19, tl=20,
          cs=22, re=23, oe=24, mi=25, te=26, tle=27, check=29)
CF = dict(ni=5, da=6, d_ar=7, d_inv=8, d_oca=9, d_ap=10, d_ocl=11, cfo=12,
          capex=14, cfi=15, div=17, ltd_repay=18, revolver=19, cff=20,
          net=22, beg=23, end=24)
DEBT = dict(rate_debt=5, rate_cash=6, min_pct=7, min_floor=8,
            ltd_beg=11, ltd_repay=12, ltd_end=13,
            min_cash=16, rev_beg=17, cash_before=18, draw=19, paydown=20,
            rev_end=21, int_exp=24, int_inc=25)

WIDTHS = {"A": 38, "B": 16, **{col_letter(i): 12 for i in range(8)}}


def _cols(n_hist):
    hist = [col_letter(i) for i in range(n_hist)]
    fc = [col_letter(n_hist + i) for i in range(N_FC)]
    return hist, fc


def _prev(c):
    """Previous data column letter (works because data cols are C..J)."""
    return chr(ord(c) - 1)


def _label(ws, row, text, bold=False, memo=False):
    set_cell(ws, f"A{row}", text, GRAY if memo else None, bold=bold)


def build_assumptions(ws, data, a, n_hist):
    hist, fc = _cols(n_hist)
    sheet_title(ws, f"{data['name']} ({data['ticker']}) — Forecast Assumptions",
                "Blue = input (seeded from historicals — challenge these). "
                "Historical columns are memo ratios per filings. "
                f"Source: FMP stable API through FY{data['years'][-1]}; "
                f"built {date.today().isoformat()}; all figures $mm.")
    _label(ws, 3, "Scenario toggle (1=Bear, 2=Base, 3=Bull)", bold=True)
    set_cell(ws, "B3", 2, BLUE, "0", fill=INPUT_HL)
    set_cell(ws, "C3", f'=CHOOSE({SELECTOR},"BEAR","BASE","BULL")', bold=True)
    dv = DataValidation(type="list", formula1='"1,2,3"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("B3")
    year_headers(ws, 4, data["years"], N_FC)

    _label(ws, 5, "Operating drivers", bold=True)
    _label(ws, 11, "Working capital", bold=True)
    _label(ws, 17, "Tax & distributions", bold=True)

    rows = [
        ("growth", "Revenue growth", FMT_PCT, a.rev_growth),
        ("gm", "Gross margin", FMT_PCT, a.gross_margin),
        ("opex", "Opex ex-D&A (% of revenue)", FMT_PCT, a.opex_pct),
        ("da", "D&A (% of revenue)", FMT_PCT, a.da_pct),
        ("capex", "Capex (% of revenue)", FMT_PCT, a.capex_pct),
        ("dso", "Days sales outstanding (DSO)", FMT_DAYS, a.dso),
        ("dio", "Days inventory outstanding (DIO)", FMT_DAYS, a.dio),
        ("dpo", "Days payables outstanding (DPO)", FMT_DAYS, a.dpo),
        ("oca", "Other current assets (% of revenue)", FMT_PCT, a.other_ca_pct),
        ("ocl", "Other current liabs (% of revenue)", FMT_PCT, a.other_cl_pct),
        ("tax", "Effective tax rate", FMT_PCT, a.tax_rate),
        ("payout", "Dividend payout (% of NI)", FMT_PCT, a.payout),
    ]
    for key, label, fmt, values in rows:
        r = ASM[key]
        if key in SCENARIO_DRIVERS:
            # Scenario-driven: live CHOOSE off the toggle; the blue inputs
            # live on the Scenarios sheet.
            _label(ws, r, label + "  [scenario-driven]")
            for c in fc:
                refs = ",".join(f"Scenarios!{c}{SCEN[s][key]}"
                                for s in ("Bear", "Base", "Bull"))
                set_cell(ws, f"{c}{r}", f"=CHOOSE({SELECTOR},{refs})",
                         None, fmt)
            continue
        _label(ws, r, label)
        for j, c in enumerate(fc):
            set_cell(ws, f"{c}{r}", round(values[j], 4), BLUE, fmt, fill=INPUT_HL)

    # Historical memo ratios (green links into the statements) for context.
    for i, c in enumerate(hist):
        if i > 0:
            set_cell(ws, f"{c}{ASM['growth']}",
                     f"=IS!{c}{IS['rev']}/IS!{_prev(c)}{IS['rev']}-1", None, FMT_PCT)
        set_cell(ws, f"{c}{ASM['gm']}",
                 f"=IF(IS!{c}{IS['rev']}=0,0,IS!{c}{IS['gp']}/IS!{c}{IS['rev']})",
                 None, FMT_PCT)
        set_cell(ws, f"{c}{ASM['opex']}",
                 f"=-IS!{c}{IS['opex']}/IS!{c}{IS['rev']}", None, FMT_PCT)
        set_cell(ws, f"{c}{ASM['da']}",
                 f"=-IS!{c}{IS['da']}/IS!{c}{IS['rev']}", None, FMT_PCT)
        set_cell(ws, f"{c}{ASM['dso']}",
                 f"=BS!{c}{BS['ar']}/IS!{c}{IS['rev']}*365", None, FMT_DAYS)
        set_cell(ws, f"{c}{ASM['dio']}",
                 f"=IF(IS!{c}{IS['cogs']}=0,0,BS!{c}{BS['inv']}"
                 f"/-IS!{c}{IS['cogs']}*365)", None, FMT_DAYS)
        set_cell(ws, f"{c}{ASM['dpo']}",
                 f"=IF(IS!{c}{IS['cogs']}=0,0,BS!{c}{BS['ap']}"
                 f"/-IS!{c}{IS['cogs']}*365)", None, FMT_DAYS)
        set_cell(ws, f"{c}{ASM['tax']}",
                 f"=IF(IS!{c}{IS['ebt']}=0,0,-IS!{c}{IS['tax']}/IS!{c}{IS['ebt']})",
                 None, FMT_PCT)
        # capex/payout ratios come from the CF statement (not rebuilt for
        # historicals) — hardcoded data, not judgment, hence blue with source.
        set_cell(ws, f"{c}{ASM['capex']}",
                 round(data["cf"]["capex"][i] / data["is"]["revenue"][i], 4),
                 BLUE, FMT_PCT)
        ni = data["is"]["ni"][i]
        set_cell(ws, f"{c}{ASM['payout']}",
                 round(data["cf"]["dividends"][i] / ni, 4) if ni > 0 else 0,
                 BLUE, FMT_PCT)
    set_widths(ws, WIDTHS)


def build_is(ws, data, n_hist):
    hist, fc = _cols(n_hist)
    d = data["is"]
    sheet_title(ws, f"{data['ticker']} — Income Statement ($mm)",
                "Historical: blue per filings (FMP). Forecast: live formulas "
                "driven by Assumptions.")
    year_headers(ws, 4, data["years"], N_FC)

    labels = [("rev", "Revenue", False), ("growth", "  % growth", True),
              ("cogs", "Cost of revenue", False), ("gp", "Gross profit", False),
              ("gm", "  % margin", True),
              ("opex", "Operating expenses (ex-D&A)", False),
              ("ebitda", "EBITDA", False), ("ebitda_m", "  % margin", True),
              ("da", "Depreciation & amortization", False),
              ("ebit", "EBIT", False),
              ("int_exp", "Interest expense", False),
              ("int_inc", "Interest income", False),
              ("other", "Other income / (expense), net", False),
              ("ebt", "Pre-tax income", False), ("tax", "Income taxes", False),
              ("etr", "  effective rate", True), ("ni", "Net income", False)]
    for key, label, memo in labels:
        _label(ws, IS[key], label, bold=key in ("ebitda", "ebit", "ni"),
               memo=memo)

    # Historicals — blue hardcodes; costs negative; subtotals as formulas.
    for i, c in enumerate(hist):
        set_cell(ws, f"{c}{IS['rev']}", d["revenue"][i], BLUE, FMT_USD)
        set_cell(ws, f"{c}{IS['cogs']}", -d["cogs"][i], BLUE, FMT_USD)
        gp = d["revenue"][i] - d["cogs"][i]
        ebitda = d["ebit"][i] + d["dna"][i]
        set_cell(ws, f"{c}{IS['opex']}", -(gp - ebitda), BLUE, FMT_USD)
        set_cell(ws, f"{c}{IS['da']}", -d["dna"][i], BLUE, FMT_USD)
        set_cell(ws, f"{c}{IS['int_exp']}", -d["int_exp"][i], BLUE, FMT_USD)
        set_cell(ws, f"{c}{IS['int_inc']}", d["int_inc"][i], BLUE, FMT_USD)
        set_cell(ws, f"{c}{IS['other']}", d["other"][i], BLUE, FMT_USD)
        set_cell(ws, f"{c}{IS['tax']}", -d["tax"][i], BLUE, FMT_USD)

    # Forecast inputs that stay analyst-editable
    for c in fc:
        set_cell(ws, f"{c}{IS['other']}", 0, BLUE, FMT_USD)

    # Formulas — identical shape across hist + forecast where applicable.
    for j, c in enumerate(fc):
        p = _prev(c)
        set_cell(ws, f"{c}{IS['rev']}",
                 f"={p}{IS['rev']}*(1+Assumptions!{c}{ASM['growth']})", None, FMT_USD)
        set_cell(ws, f"{c}{IS['cogs']}",
                 f"=-{c}{IS['rev']}*(1-Assumptions!{c}{ASM['gm']})", None, FMT_USD)
        set_cell(ws, f"{c}{IS['opex']}",
                 f"=-{c}{IS['rev']}*Assumptions!{c}{ASM['opex']}", None, FMT_USD)
        set_cell(ws, f"{c}{IS['da']}",
                 f"=-{c}{IS['rev']}*Assumptions!{c}{ASM['da']}", None, FMT_USD)
        set_cell(ws, f"{c}{IS['int_exp']}", f"=-Debt!{c}{DEBT['int_exp']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{IS['int_inc']}", f"=Debt!{c}{DEBT['int_inc']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{IS['tax']}",
                 f"=-{c}{IS['ebt']}*Assumptions!{c}{ASM['tax']}", None, FMT_USD)

    for i, c in enumerate(hist + fc):
        if i > 0:
            set_cell(ws, f"{c}{IS['growth']}",
                     f"={c}{IS['rev']}/{_prev(c)}{IS['rev']}-1", None, FMT_PCT)
        set_cell(ws, f"{c}{IS['gp']}", f"={c}{IS['rev']}+{c}{IS['cogs']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{IS['gm']}", f"={c}{IS['gp']}/{c}{IS['rev']}",
                 None, FMT_PCT)
        set_cell(ws, f"{c}{IS['ebitda']}", f"={c}{IS['gp']}+{c}{IS['opex']}",
                 None, FMT_USD, bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{IS['ebitda_m']}",
                 f"={c}{IS['ebitda']}/{c}{IS['rev']}", None, FMT_PCT)
        set_cell(ws, f"{c}{IS['ebit']}", f"={c}{IS['ebitda']}+{c}{IS['da']}",
                 None, FMT_USD, bold=True)
        set_cell(ws, f"{c}{IS['ebt']}",
                 f"=SUM({c}{IS['ebit']}:{c}{IS['other']})", None, FMT_USD)
        set_cell(ws, f"{c}{IS['etr']}",
                 f"=IF({c}{IS['ebt']}=0,0,-{c}{IS['tax']}/{c}{IS['ebt']})",
                 None, FMT_PCT)
        set_cell(ws, f"{c}{IS['ni']}", f"={c}{IS['ebt']}+{c}{IS['tax']}",
                 None, FMT_USD, bold=True, border=THIN_TOP)
    set_widths(ws, WIDTHS)


def build_bs(ws, data, n_hist):
    hist, fc = _cols(n_hist)
    d = data["bs"]
    sheet_title(ws, f"{data['ticker']} — Balance Sheet ($mm)",
                '"Other" lines are explicit plugs from reported subtotals so '
                "historicals tie exactly; held constant or revenue-scaled in "
                "the forecast.")
    year_headers(ws, 4, data["years"], N_FC)

    labels = [("cash", "Cash & ST investments"), ("ar", "Accounts receivable"),
              ("inv", "Inventory"), ("oca", "Other current assets"),
              ("tca", "Total current assets"), ("ppe", "Net PP&E"),
              ("gwi", "Goodwill & intangibles"),
              ("onca", "Other non-current assets"), ("ta", "Total assets"),
              ("ap", "Accounts payable"), ("ocl", "Other current liabilities"),
              ("rev", "Revolver"), ("ltd", "Long-term debt (incl. current)"),
              ("oncl", "Other non-current liabilities"),
              ("tl", "Total liabilities"), ("cs", "Common stock & APIC"),
              ("re", "Retained earnings"),
              ("oe", "Other equity (AOCI, treasury)"),
              ("mi", "Minority interest"), ("te", "Total equity"),
              ("tle", "Total liabilities & equity"),
              ("check", "Balance check: TA − TL&E (must be 0)")]
    for key, label in labels:
        _label(ws, BS[key], label,
               bold=key in ("tca", "ta", "tl", "te", "tle"), memo=key == "check")

    hist_map = {"cash": "cash", "ar": "ar", "inv": "inv", "oca": "other_ca",
                "ppe": "ppe", "gwi": "gw_intan", "onca": "other_nca",
                "ap": "ap", "ocl": "other_cl", "ltd": "debt",
                "oncl": "other_ncl", "cs": "cs_apic", "re": "re",
                "oe": "other_eq", "mi": "minority"}
    for i, c in enumerate(hist):
        for row_key, data_key in hist_map.items():
            set_cell(ws, f"{c}{BS[row_key]}", d[data_key][i], BLUE, FMT_USD)
        set_cell(ws, f"{c}{BS['rev']}", 0, BLUE, FMT_USD)  # modeled line; starts at 0

    for j, c in enumerate(fc):
        p = _prev(c)
        set_cell(ws, f"{c}{BS['cash']}", f"=CF!{c}{CF['end']}", None, FMT_USD)
        set_cell(ws, f"{c}{BS['ar']}",
                 f"=Assumptions!{c}{ASM['dso']}/365*IS!{c}{IS['rev']}", None, FMT_USD)
        set_cell(ws, f"{c}{BS['inv']}",
                 f"=Assumptions!{c}{ASM['dio']}/365*-IS!{c}{IS['cogs']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{BS['oca']}",
                 f"=Assumptions!{c}{ASM['oca']}*IS!{c}{IS['rev']}", None, FMT_USD)
        set_cell(ws, f"{c}{BS['ppe']}",
                 f"={p}{BS['ppe']}-CF!{c}{CF['capex']}+IS!{c}{IS['da']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{BS['gwi']}", f"={p}{BS['gwi']}", None, FMT_USD)
        set_cell(ws, f"{c}{BS['onca']}", f"={p}{BS['onca']}", None, FMT_USD)
        set_cell(ws, f"{c}{BS['ap']}",
                 f"=Assumptions!{c}{ASM['dpo']}/365*-IS!{c}{IS['cogs']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{BS['ocl']}",
                 f"=Assumptions!{c}{ASM['ocl']}*IS!{c}{IS['rev']}", None, FMT_USD)
        set_cell(ws, f"{c}{BS['rev']}", f"=Debt!{c}{DEBT['rev_end']}", None, FMT_USD)
        set_cell(ws, f"{c}{BS['ltd']}", f"=Debt!{c}{DEBT['ltd_end']}", None, FMT_USD)
        set_cell(ws, f"{c}{BS['oncl']}", f"={p}{BS['oncl']}", None, FMT_USD)
        set_cell(ws, f"{c}{BS['cs']}", f"={p}{BS['cs']}", None, FMT_USD)
        set_cell(ws, f"{c}{BS['re']}",
                 f"={p}{BS['re']}+IS!{c}{IS['ni']}+CF!{c}{CF['div']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{BS['oe']}", f"={p}{BS['oe']}", None, FMT_USD)
        set_cell(ws, f"{c}{BS['mi']}", f"={p}{BS['mi']}", None, FMT_USD)

    for c in hist + fc:
        set_cell(ws, f"{c}{BS['tca']}",
                 f"=SUM({c}{BS['cash']}:{c}{BS['oca']})", None, FMT_USD,
                 bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{BS['ta']}",
                 f"={c}{BS['tca']}+{c}{BS['ppe']}+{c}{BS['gwi']}+{c}{BS['onca']}",
                 None, FMT_USD, bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{BS['tl']}",
                 f"=SUM({c}{BS['ap']}:{c}{BS['oncl']})", None, FMT_USD,
                 bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{BS['te']}",
                 f"=SUM({c}{BS['cs']}:{c}{BS['mi']})", None, FMT_USD,
                 bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{BS['tle']}",
                 f"={c}{BS['tl']}+{c}{BS['te']}", None, FMT_USD,
                 bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{BS['check']}",
                 f"={c}{BS['ta']}-{c}{BS['tle']}", None, FMT_USD)
    set_widths(ws, WIDTHS)


def build_cf(ws, data, n_hist):
    hist, fc = _cols(n_hist)
    sheet_title(ws, f"{data['ticker']} — Cash Flow Statement ($mm, forecast)",
                "Forecast build only — historical cash flows per filings are "
                "not restated. Ending cash flows to the balance sheet.")
    year_headers(ws, 4, data["years"], N_FC)

    labels = [("ni", "Net income"), ("da", "(+) D&A"),
              ("d_ar", "(Inc)/dec accounts receivable"),
              ("d_inv", "(Inc)/dec inventory"),
              ("d_oca", "(Inc)/dec other current assets"),
              ("d_ap", "Inc/(dec) accounts payable"),
              ("d_ocl", "Inc/(dec) other current liabilities"),
              ("cfo", "Cash from operations"), ("capex", "Capital expenditures"),
              ("cfi", "Cash from investing"), ("div", "Dividends paid"),
              ("ltd_repay", "LTD scheduled repayment"),
              ("revolver", "Revolver draw / (paydown)"),
              ("cff", "Cash from financing"), ("net", "Net change in cash"),
              ("beg", "Beginning cash"), ("end", "Ending cash")]
    for key, label in labels:
        _label(ws, CF[key], label, bold=key in ("cfo", "cfi", "cff", "net", "end"))

    for j, c in enumerate(fc):
        p = _prev(c)
        set_cell(ws, f"{c}{CF['ni']}", f"=IS!{c}{IS['ni']}", None, FMT_USD)
        set_cell(ws, f"{c}{CF['da']}", f"=-IS!{c}{IS['da']}", None, FMT_USD)
        for key, bs_key in (("d_ar", "ar"), ("d_inv", "inv"), ("d_oca", "oca")):
            set_cell(ws, f"{c}{CF[key]}",
                     f"=-(BS!{c}{BS[bs_key]}-BS!{p}{BS[bs_key]})", None, FMT_USD)
        for key, bs_key in (("d_ap", "ap"), ("d_ocl", "ocl")):
            set_cell(ws, f"{c}{CF[key]}",
                     f"=BS!{c}{BS[bs_key]}-BS!{p}{BS[bs_key]}", None, FMT_USD)
        set_cell(ws, f"{c}{CF['cfo']}",
                 f"=SUM({c}{CF['ni']}:{c}{CF['d_ocl']})", None, FMT_USD,
                 bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{CF['capex']}",
                 f"=-IS!{c}{IS['rev']}*Assumptions!{c}{ASM['capex']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{CF['cfi']}", f"={c}{CF['capex']}", None, FMT_USD,
                 bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{CF['div']}",
                 f"=-MAX(0,IS!{c}{IS['ni']})*Assumptions!{c}{ASM['payout']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{CF['ltd_repay']}", f"=-Debt!{c}{DEBT['ltd_repay']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{CF['revolver']}",
                 f"=Debt!{c}{DEBT['draw']}+Debt!{c}{DEBT['paydown']}",
                 None, FMT_USD)
        set_cell(ws, f"{c}{CF['cff']}",
                 f"=SUM({c}{CF['div']}:{c}{CF['revolver']})", None, FMT_USD,
                 bold=True, border=THIN_TOP)
        set_cell(ws, f"{c}{CF['net']}",
                 f"={c}{CF['cfo']}+{c}{CF['cfi']}+{c}{CF['cff']}", None, FMT_USD,
                 bold=True)
        beg = f"=BS!{p}{BS['cash']}" if j == 0 else f"={p}{CF['end']}"
        set_cell(ws, f"{c}{CF['beg']}", beg, None, FMT_USD)
        set_cell(ws, f"{c}{CF['end']}", f"={c}{CF['beg']}+{c}{CF['net']}",
                 None, FMT_USD, bold=True, border=THIN_TOP)
    set_widths(ws, WIDTHS)


def build_debt(ws, data, a, n_hist):
    hist, fc = _cols(n_hist)
    sheet_title(ws, f"{data['ticker']} — Debt Schedule & Revolver ($mm)",
                "Interest on BEGINNING-of-period balances (no circularity — "
                "see README). Revolver draws to the minimum cash target and "
                "sweeps excess cash back down.")
    year_headers(ws, 4, data["years"], N_FC)

    scalars = [("rate_debt", "Interest rate on debt", a.rate_debt, FMT_PCT),
               ("rate_cash", "Yield on cash", a.rate_cash, FMT_PCT),
               ("min_pct", "Minimum cash — % of revenue", a.min_cash_pct, FMT_PCT),
               ("min_floor", "Minimum cash — floor ($mm)", a.min_cash_floor,
                FMT_USD)]
    for key, label, val, fmt in scalars:
        _label(ws, DEBT[key], label)
        set_cell(ws, f"B{DEBT[key]}", val, BLUE, fmt, fill=INPUT_HL)

    _label(ws, 10, "Long-term debt", bold=True)
    _label(ws, DEBT["ltd_beg"], "Beginning balance")
    _label(ws, DEBT["ltd_repay"], "Scheduled repayment (input)")
    _label(ws, DEBT["ltd_end"], "Ending balance")
    _label(ws, 15, "Revolver & cash sweep", bold=True)
    _label(ws, DEBT["min_cash"], "Minimum cash required")
    _label(ws, DEBT["rev_beg"], "Revolver — beginning")
    _label(ws, DEBT["cash_before"], "Cash before revolver")
    _label(ws, DEBT["draw"], "Revolver draw")
    _label(ws, DEBT["paydown"], "Revolver (paydown)")
    _label(ws, DEBT["rev_end"], "Revolver — ending")
    _label(ws, 23, "Interest (beginning balances)", bold=True)
    _label(ws, DEBT["int_exp"], "Interest expense")
    _label(ws, DEBT["int_inc"], "Interest income")

    last_hist = hist[-1]
    for j, c in enumerate(fc):
        p = _prev(c)
        ltd_beg = f"=BS!{last_hist}{BS['ltd']}" if j == 0 else f"={p}{DEBT['ltd_end']}"
        set_cell(ws, f"{c}{DEBT['ltd_beg']}", ltd_beg, None, FMT_USD)
        set_cell(ws, f"{c}{DEBT['ltd_repay']}", 0, BLUE, FMT_USD)
        set_cell(ws, f"{c}{DEBT['ltd_end']}",
                 f"={c}{DEBT['ltd_beg']}-{c}{DEBT['ltd_repay']}", None, FMT_USD)
        set_cell(ws, f"{c}{DEBT['min_cash']}",
                 f"=MAX($B${DEBT['min_floor']},$B${DEBT['min_pct']}"
                 f"*IS!{c}{IS['rev']})", None, FMT_USD)
        rev_beg = f"=BS!{last_hist}{BS['rev']}" if j == 0 else f"={p}{DEBT['rev_end']}"
        set_cell(ws, f"{c}{DEBT['rev_beg']}", rev_beg, None, FMT_USD)
        set_cell(ws, f"{c}{DEBT['cash_before']}",
                 f"=CF!{c}{CF['beg']}+CF!{c}{CF['cfo']}+CF!{c}{CF['cfi']}"
                 f"+CF!{c}{CF['div']}+CF!{c}{CF['ltd_repay']}", None, FMT_USD)
        set_cell(ws, f"{c}{DEBT['draw']}",
                 f"=MAX(0,{c}{DEBT['min_cash']}-{c}{DEBT['cash_before']})",
                 None, FMT_USD)
        set_cell(ws, f"{c}{DEBT['paydown']}",
                 f"=-MIN({c}{DEBT['rev_beg']},"
                 f"MAX(0,{c}{DEBT['cash_before']}-{c}{DEBT['min_cash']}))",
                 None, FMT_USD)
        set_cell(ws, f"{c}{DEBT['rev_end']}",
                 f"={c}{DEBT['rev_beg']}+{c}{DEBT['draw']}+{c}{DEBT['paydown']}",
                 None, FMT_USD, border=THIN_TOP)
        set_cell(ws, f"{c}{DEBT['int_exp']}",
                 f"=$B${DEBT['rate_debt']}*({c}{DEBT['ltd_beg']}"
                 f"+{c}{DEBT['rev_beg']})", None, FMT_USD)
        set_cell(ws, f"{c}{DEBT['int_inc']}",
                 f"=$B${DEBT['rate_cash']}*CF!{c}{CF['beg']}", None, FMT_USD)
    set_widths(ws, WIDTHS)


def build_scenarios(ws, data, a, n_hist):
    hist, fc = _cols(n_hist)
    sheet_title(ws, f"{data['ticker']} — Scenario Drivers",
                "Three blocks of blue inputs; the Assumptions sheet CHOOSEs a "
                "block via the toggle. Bear/Bull are mechanical seeds off "
                "Base — overwrite them with a view.")
    year_headers(ws, 4, data["years"], N_FC)

    scen = seed_scenarios(a)
    driver_labels = dict(growth="Revenue growth", gm="Gross margin",
                         opex="Opex ex-D&A (% of revenue)",
                         capex="Capex (% of revenue)")
    seeds = {"Bear": f"seed: growth {BEAR['growth']:+.0%}, GM {BEAR['gm']:+.1%}, "
                     f"opex {BEAR['opex']:+.1%}, capex {BEAR['capex']:+.1%}",
             "Base": "seed: derived from historicals",
             "Bull": f"seed: growth {BULL['growth']:+.0%}, GM {BULL['gm']:+.1%}, "
                     f"opex {BULL['opex']:+.1%}, capex {BULL['capex']:+.1%}"}
    for name in ("Bear", "Base", "Bull"):
        block = SCEN[name]
        header_r = min(block.values()) - 1
        _label(ws, header_r, f"{name} case", bold=True)
        set_cell(ws, f"B{header_r}", seeds[name], GRAY)
        for key, r in block.items():
            _label(ws, r, driver_labels[key])
            for j, c in enumerate(fc):
                set_cell(ws, f"{c}{r}", round(scen[name][key][j], 4), BLUE,
                         FMT_PCT, fill=INPUT_HL)
    set_widths(ws, WIDTHS)


def build_checks(ws, data, n_hist):
    hist, fc = _cols(n_hist)
    sheet_title(ws, f"{data['ticker']} — Model Integrity Checks",
                "Every value on this sheet must be zero; B9 must read PASS. "
                "The automated recalc gate asserts this.")
    year_headers(ws, 4, data["years"], N_FC)

    _label(ws, 5, "Balance sheet: TA − TL&E")
    _label(ws, 6, "Revolver non-negative (0 = ok)")
    _label(ws, 7, "Cash ≥ minimum target (0 = ok)")
    for c in hist + fc:
        set_cell(ws, f"{c}5", f"=BS!{c}{BS['ta']}-BS!{c}{BS['tle']}",
                 None, "0.000")
    for c in fc:
        set_cell(ws, f"{c}6",
                 f"=IF(Debt!{c}{DEBT['rev_end']}>=-0.001,0,1)", None, "0")
        set_cell(ws, f"{c}7",
                 f"=IF(CF!{c}{CF['end']}>=Debt!{c}{DEBT['min_cash']}-0.01,0,1)",
                 None, "0")
    first, last = (hist + fc)[0], (hist + fc)[-1]
    _label(ws, 9, "ALL CHECKS", bold=True)
    cell = set_cell(ws, "B9",
                    f'=IF(SUMPRODUCT(ABS({first}5:{last}5))'
                    f'+SUM({fc[0]}6:{fc[-1]}7)<0.005,"PASS","FAIL")')
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    set_widths(ws, WIDTHS)


def build_model(data, assumptions, out_path):
    n_hist = data["n_hist"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    build_assumptions(ws, data, assumptions, n_hist)
    build_scenarios(wb.create_sheet("Scenarios"), data, assumptions, n_hist)
    build_is(wb.create_sheet("IS"), data, n_hist)
    build_bs(wb.create_sheet("BS"), data, n_hist)
    build_cf(wb.create_sheet("CF"), data, n_hist)
    build_debt(wb.create_sheet("Debt"), data, assumptions, n_hist)
    build_checks(wb.create_sheet("Checks"), data, n_hist)
    wb.save(out_path)
    return out_path
