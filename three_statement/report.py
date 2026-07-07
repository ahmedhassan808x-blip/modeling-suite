"""
Presentation & export layer for the three-statement model.

    python3 -m three_statement.report AAPL_3stmt.xlsx

Produces, next to the workbook: {stem}_deck.pptx, {stem}_memo.docx,
{stem}_deck.pdf. Every number is extracted from the RECALCULATED workbook
(LibreOffice), so the deck cannot disagree with the model — including the
scenario slide, which flips the in-workbook toggle and re-runs the engine
for each case.
"""

import re
import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from shared import charts  # noqa: E402
from shared.exports.docx_memo import Memo  # noqa: E402
from shared.exports.pdf import to_pdf  # noqa: E402
from shared.exports.pptx_deck import Deck  # noqa: E402
from shared.recalc import recalculate  # noqa: E402
from three_statement.model_builder import ASM, BS, CF, DEBT, IS  # noqa: E402

COLS = list("CDEFGHIJ")   # 3 hist + 5 forecast
FC = list("FGHIJ")
N_HIST = 3

LIMITATIONS = [
    "No share buybacks or issuance modeled — cash-rich companies accumulate "
    "large cash balances in later years, and interest income on that cash "
    "flatters late-year margins.",
    "Interest accrues on beginning-of-period balances (avoids circular "
    "references; slightly understates interest when balances grow).",
    "Goodwill, other assets/liabilities and equity lines held flat; single "
    "debt tranche plus revolver; flat effective tax rate with no NOL tracking.",
    "Forecast drivers are seeded mechanically from 3 years of history — they "
    "are starting points for judgment, not predictions.",
]


def _probes():
    p = ["Checks!B9"]
    for c in COLS:
        p += [f"Assumptions!{c}4", f"IS!{c}{IS['rev']}",
              f"IS!{c}{IS['ebitda']}", f"IS!{c}{IS['ebitda_m']}",
              f"IS!{c}{IS['ni']}", f"BS!{c}{BS['cash']}",
              f"BS!{c}{BS['ltd']}", f"BS!{c}{BS['rev']}"]
    for c in FC:
        p += [f"Debt!{c}{DEBT['rev_end']}", f"Debt!{c}{DEBT['min_cash']}"]
        for key in ("growth", "gm", "opex", "capex", "da", "tax", "payout"):
            p.append(f"Assumptions!{c}{ASM[key]}")
    return p


def extract(xlsx) -> dict:
    """Recalculate the workbook and pull everything the report needs."""
    res = recalculate(xlsx, probe_cells=_probes())
    if not res.ok:
        raise RuntimeError(f"{xlsx}: refusing to report on a workbook with "
                           f"formula errors.\n{res.summary()}")
    v = res.values
    wb = load_workbook(xlsx)
    title = wb["Assumptions"]["A1"].value or ""
    m = re.match(r"^(?P<name>.+?) \((?P<ticker>[A-Z.\-]+)\)", title)
    name, ticker = (m["name"], m["ticker"]) if m else (title, "?")

    def row(sheet, r, cols=COLS):
        return [v[f"{sheet}!{c}{r}"] for c in cols]

    years = [str(v[f"Assumptions!{c}4"]) for c in COLS]
    drivers = {key: [v[f"Assumptions!{c}{ASM[key]}"] for c in FC]
               for key in ("growth", "gm", "opex", "capex", "da", "tax",
                           "payout")}
    return dict(
        name=name, ticker=ticker, years=years, checks=v["Checks!B9"],
        revenue=row("IS", IS["rev"]), ebitda=row("IS", IS["ebitda"]),
        ebitda_m=row("IS", IS["ebitda_m"]), ni=row("IS", IS["ni"]),
        cash=row("BS", BS["cash"]), ltd=row("BS", BS["ltd"]),
        revolver=row("BS", BS["rev"]),
        rev_end=[v[f"Debt!{c}{DEBT['rev_end']}"] for c in FC],
        min_cash=[v[f"Debt!{c}{DEBT['min_cash']}"] for c in FC],
        drivers=drivers,
    )


def scenario_runs(xlsx, extra_probes=None) -> dict:
    """Flip the in-workbook toggle per scenario and recalculate each case."""
    out = {}
    probes = [f"IS!{c}{IS['rev']}" for c in FC] + \
        [f"IS!J{IS['ni']}", "Checks!B9"] + (extra_probes or [])
    with tempfile.TemporaryDirectory() as tmp:
        for sel, label in ((1, "Bear"), (2, "Base"), (3, "Bull")):
            p = Path(tmp) / f"s{sel}.xlsx"
            shutil.copy(xlsx, p)
            wb = load_workbook(p)
            wb["Assumptions"]["B3"] = sel
            wb.save(p)
            res = recalculate(p, probe_cells=probes)
            assert res.values["Checks!B9"] == "PASS", f"{label} failed checks"
            out[label] = res.values
    return out


# ---- report assembly ------------------------------------------------------

def _fmt(v, kind="usd"):
    if kind == "usd":
        return f"${v:,.0f}"
    if kind == "pct":
        return f"{v:.1%}"
    if kind == "ps":
        return f"${v:,.2f}"
    return str(v)


def _driver_table(d, years_fc):
    rows = [["Driver ($mm basis)"] + years_fc]
    labels = [("growth", "Revenue growth", "pct"), ("gm", "Gross margin", "pct"),
              ("opex", "Opex ex-D&A (% rev)", "pct"),
              ("capex", "Capex (% rev)", "pct"), ("da", "D&A (% rev)", "pct"),
              ("tax", "Tax rate", "pct"), ("payout", "Dividend payout", "pct")]
    for key, label, kind in labels:
        rows.append([label] + [_fmt(x, kind) for x in d["drivers"][key]])
    return rows


def build_reports(xlsx, outdir=None):
    xlsx = Path(xlsx)
    outdir = Path(outdir) if outdir else xlsx.parent
    d = extract(xlsx)
    scen = scenario_runs(xlsx)
    years_fc = d["years"][N_HIST:]
    hist_n = N_HIST
    debt_total = [l + r for l, r in zip(d["ltd"], d["revolver"])]

    tmp = Path(tempfile.mkdtemp())
    png_rev = charts.revenue_ebitda(d["years"], d["revenue"], d["ebitda_m"],
                                    hist_n, tmp / "rev.png")
    paths = {lbl: [d["revenue"][hist_n - 1]] +
             [scen[lbl][f"IS!{c}{IS['rev']}"] for c in FC]
             for lbl in ("Bear", "Base", "Bull")}
    png_scen = charts.scenario_paths([d["years"][hist_n - 1]] + years_fc,
                                     paths, tmp / "scen.png")
    png_cash = charts.cash_and_debt(d["years"], d["cash"], debt_total,
                                    tmp / "cash.png")

    rev, ni = d["revenue"], d["ni"]
    cagr_h = (rev[hist_n - 1] / rev[0]) ** (1 / (hist_n - 1)) - 1
    cagr_f = (rev[-1] / rev[hist_n - 1]) ** (1 / len(years_fc)) - 1
    max_drawn = max(d["rev_end"])
    summary_bullets = [
        f"Revenue {_fmt(rev[0])}mm → {_fmt(rev[hist_n - 1])}mm over the last "
        f"{hist_n} fiscal years ({cagr_h:.1%} CAGR); forecast to "
        f"{_fmt(rev[-1])}mm by {d['years'][-1]} ({cagr_f:.1%} CAGR, growth "
        "fading toward terminal).",
        f"EBITDA margin {d['ebitda_m'][hist_n - 1]:.1%} last actual; "
        f"{d['ebitda_m'][-1]:.1%} in the terminal forecast year.",
        f"Net income {_fmt(ni[hist_n - 1])}mm → {_fmt(ni[-1])}mm by "
        f"{d['years'][-1]}; ending cash {_fmt(d['cash'][-1])}mm.",
        ("Revolver undrawn across the forecast — the model self-funds above "
         "its minimum cash target." if max_drawn <= 0.01 else
         f"Revolver peaks at {_fmt(max_drawn)}mm drawn — the cash sweep is "
         "doing real work."),
        f"Balance sheet check: {d['checks']} in every historical and forecast "
        "year, verified by automated recalculation.",
    ]
    scen_table = [["Scenario", f"Revenue {d['years'][-1]}",
                   f"Net income {d['years'][-1]}"]]
    for lbl in ("Bear", "Base", "Bull"):
        scen_table.append([lbl, _fmt(scen[lbl][f"IS!J{IS['rev']}"]) + "mm",
                           _fmt(scen[lbl][f"IS!J{IS['ni']}"]) + "mm"])

    footer = (f"{d['ticker']} — three-statement operating model | "
              "modeling-suite | all figures $mm unless noted")

    # ---- deck ----
    deck = Deck(footer)
    deck.title_slide(f"{d['name']} — Operating Model",
                     f"Three-statement forecast, FY{d['years'][0][2:]} – "
                     f"{d['years'][-1]}",
                     "Live Excel model: blue = input, black = formula, "
                     "green = cross-sheet link")
    s = deck.content_slide("Executive summary")
    deck.add_bullets(s, summary_bullets, 0.5, 1.3, 5.4, 5.3, size=12)
    deck.add_image(s, png_rev, 6.1, 1.5, 6.8)

    s = deck.content_slide("Key assumptions — Base case")
    deck.add_table(s, _driver_table(d, years_fc), 0.5, 1.4, 8.6,
                   first_col_w=2.6)
    deck.add_bullets(s, [
        "Every driver seeded from the company's own 3-year history, then "
        "exposed as a visible input — nothing judgment-based is buried in "
        "formulas.",
        "Working capital driven by DSO / DIO / DPO days, not flat "
        "percentages.",
    ], 0.5, 4.4, 11.5, 2.0, size=11)

    s = deck.content_slide("Scenario analysis — flip one cell, the model "
                           "reprices")
    deck.add_image(s, png_scen, 0.5, 1.4, 7.2)
    deck.add_table(s, scen_table, 8.1, 1.7, 4.6)
    deck.add_bullets(s, [
        "Bear/Bull are mechanical seeds off Base (growth ∓300bps, margin "
        "∓100–150bps) — meant to be overwritten with a view.",
    ], 8.1, 3.6, 4.6, 2.4, size=10)

    s = deck.content_slide("Liquidity & capital structure")
    deck.add_image(s, png_cash, 0.5, 1.4, 7.2)
    deck.add_bullets(s, [
        f"Minimum cash target: greater of a fixed floor or % of revenue "
        f"({_fmt(d['min_cash'][0])}mm in Y1).",
        ("Revolver never draws in the base case." if max_drawn <= 0.01 else
         f"Peak revolver usage {_fmt(max_drawn)}mm."),
        "Interest on beginning-of-period balances — a deliberate, documented "
        "convention that keeps the model free of circular references.",
    ], 8.0, 1.6, 4.8, 4.5, size=11)

    s = deck.content_slide("Limitations — read before relying on the numbers")
    deck.add_bullets(s, LIMITATIONS, 0.5, 1.4, 12.0, 5.0, size=12)
    pptx_path = deck.save(outdir / f"{xlsx.stem}_deck.pptx")

    # ---- memo ----
    memo = Memo(f"{d['name']} ({d['ticker']}) — Operating Model Memo",
                "Three-statement forecast summary | modeling-suite")
    memo.heading("Executive summary")
    memo.bullets(summary_bullets)
    memo.image(png_rev)
    memo.heading("Key assumptions (Base case)")
    memo.table(_driver_table(d, years_fc))
    memo.heading("Scenarios")
    memo.image(png_scen)
    memo.table(scen_table)
    memo.heading("Liquidity")
    memo.image(png_cash)
    memo.heading("Limitations")
    memo.bullets(LIMITATIONS)
    memo.para("Generated by modeling-suite from the recalculated workbook — "
              "figures cannot diverge from the model.", small=True)
    docx_path = memo.save(outdir / f"{xlsx.stem}_memo.docx")

    pdf_path = to_pdf(pptx_path, outdir)
    shutil.rmtree(tmp, ignore_errors=True)
    return dict(pptx=pptx_path, docx=docx_path, pdf=pdf_path)


def main():
    import argparse
    ap = argparse.ArgumentParser(description="Build deck/memo/PDF from a "
                                 "3-statement workbook.")
    ap.add_argument("workbook")
    ap.add_argument("--outdir")
    args = ap.parse_args()
    out = build_reports(args.workbook, args.outdir)
    for k, p in out.items():
        print(f"{k}: {p}")


if __name__ == "__main__":
    main()
